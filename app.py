import os, sqlite3, traceback, secrets, urllib.parse
from datetime import datetime, timezone
from functools import wraps
from flask import Flask, jsonify, request, g, send_from_directory, session, redirect
from dotenv import load_dotenv
from apscheduler.schedulers.background import BackgroundScheduler
import requests
from cryptography.fernet import Fernet, InvalidToken

load_dotenv()

# Patch Werkzeug Response.set_cookie to drop 'partitioned' kwarg if this build doesn't support it.
# Needed when Flask>=3.1 is paired with an older Werkzeug that lacks the partitioned parameter.
import inspect as _inspect
from werkzeug.wrappers import Response as _WResponse
if 'partitioned' not in _inspect.signature(_WResponse.set_cookie).parameters:
    _orig_set_cookie = _WResponse.set_cookie
    def _patched_set_cookie(self, *args, **kwargs):
        kwargs.pop('partitioned', None)
        return _orig_set_cookie(self, *args, **kwargs)
    _WResponse.set_cookie = _patched_set_cookie

BASE_DIR   = os.path.dirname(os.path.abspath(__file__))
DB_PATH    = os.path.join(BASE_DIR, "closeapp.db")
STATIC_DIR = os.path.join(BASE_DIR, "static")

_ENV_PATH = os.path.join(BASE_DIR, ".env")

def _persist_env(key, value):
    """Append KEY=VALUE to .env so subsequent runs reuse the same secret.
    Silent no-op on filesystem errors — the generated value still works
    for this process; next run will just generate a new one."""
    try:
        with open(_ENV_PATH, "a", encoding="utf-8") as f:
            f.write(f"\n{key}={value}\n")
        print(f"[security] generated {key} and saved to {_ENV_PATH}", flush=True)
    except OSError as e:
        print(f"[security] generated {key} in-memory (could not write {_ENV_PATH}: {e})", flush=True)

SECRET_KEY = os.getenv("SECRET_KEY", "").strip()
if not SECRET_KEY:
    SECRET_KEY = secrets.token_hex(32)
    _persist_env("SECRET_KEY", SECRET_KEY)

# Fernet-encrypts QB access/refresh tokens at rest. Auto-generates a key on
# first run and writes it to .env. Legacy plaintext token rows are read back
# transparently by decrypt_token(), so rolling this out does not break an
# existing qb_tokens table.
_TOKEN_KEY = os.getenv("TOKEN_ENCRYPTION_KEY", "").strip()
if not _TOKEN_KEY:
    _TOKEN_KEY = Fernet.generate_key().decode()
    _persist_env("TOKEN_ENCRYPTION_KEY", _TOKEN_KEY)
_FERNET = Fernet(_TOKEN_KEY.encode())

# Cross-origin requests are denied unless the Origin matches this whitelist.
# Defaults cover the local Flask server; add more via ALLOWED_ORIGINS=a,b,c.
_DEFAULT_ORIGINS = {"http://127.0.0.1:5000", "http://localhost:5000"}
ALLOWED_ORIGINS = _DEFAULT_ORIGINS | {
    o.strip() for o in os.getenv("ALLOWED_ORIGINS", "").split(",") if o.strip()
}

app = Flask(__name__, static_folder=None)
app.secret_key = SECRET_KEY
app.config.update(
    SESSION_COOKIE_HTTPONLY=True,
    SESSION_COOKIE_SAMESITE="Lax",
    SESSION_COOKIE_SECURE=os.getenv("SESSION_COOKIE_SECURE", "0") == "1",
)

@app.after_request
def add_cors(response):
    origin = request.headers.get("Origin")
    if origin and origin in ALLOWED_ORIGINS:
        response.headers["Access-Control-Allow-Origin"] = origin
        response.headers["Vary"] = "Origin"
        response.headers["Access-Control-Allow-Methods"] = "GET,POST,PATCH,DELETE,OPTIONS"
        response.headers["Access-Control-Allow-Headers"] = "Content-Type,X-CSRF-Token"
        response.headers["Access-Control-Allow-Credentials"] = "true"
    return response

def get_db():
    if "db" not in g:
        g.db = sqlite3.connect(DB_PATH)
        g.db.row_factory = sqlite3.Row
        g.db.execute("PRAGMA foreign_keys=ON")
    return g.db

@app.teardown_appcontext
def close_db(exc):
    db = g.pop("db", None)
    if db: db.close()

def q(sql, params=()):
    return get_db().execute(sql, params).fetchall()

def q1(sql, params=()):
    return get_db().execute(sql, params).fetchone()

def run(sql, params=()):
    db = get_db()
    cur = db.execute(sql, params)
    db.commit()
    return cur

def rows_to_list(rows):
    return [dict(r) for r in rows]

def err(msg, code=400):
    return jsonify({"error": msg}), code

def safe_update(table, pk_col, allowed_cols, updates, pk_value):
    """UPDATE <table> SET k=? ... WHERE <pk_col>=? with every column name
    validated against a whitelist. Prevents f-string SQL injection if an
    upstream filter is ever loosened."""
    if not table.isidentifier() or not pk_col.isidentifier():
        raise ValueError("Invalid table/column identifier")
    if not updates:
        raise ValueError("No updates provided")
    invalid = set(updates) - set(allowed_cols)
    if invalid:
        raise ValueError(f"Disallowed column(s): {sorted(invalid)}")
    for k in updates:
        if not k.isidentifier():
            raise ValueError(f"Invalid column identifier: {k}")
    set_clause = ", ".join(f"{k}=?" for k in updates)
    return run(
        f"UPDATE {table} SET {set_clause} WHERE {pk_col}=?",
        list(updates.values()) + [pk_value],
    )

def encrypt_token(plain):
    if plain is None or not _FERNET:
        return plain
    return _FERNET.encrypt(plain.encode()).decode()

def decrypt_token(stored):
    if stored is None or not _FERNET:
        return stored
    try:
        return _FERNET.decrypt(stored.encode()).decode()
    except InvalidToken:
        return stored  # legacy plaintext row — leave alone

def csrf_protect(f):
    """Reject POST/PATCH/PUT/DELETE without a valid X-CSRF-Token header.
    The token is seeded by /api/auth/me and lives in the session cookie."""
    @wraps(f)
    def decorated(*args, **kwargs):
        if request.method in ("POST", "PATCH", "PUT", "DELETE"):
            sent = request.headers.get("X-CSRF-Token", "")
            expected = session.get("csrf_token", "")
            if not expected or not sent or not secrets.compare_digest(sent, expected):
                return jsonify({"error": "Invalid CSRF token"}), 403
        return f(*args, **kwargs)
    return decorated

# ── Auth disabled ─────────────────────────────────────────────────────────────
# Login screen intentionally removed. The app runs as the seeded admin
# (user id 1 — "Joe G.") for every request. See AGENTS.md §1 before touching.

DEFAULT_USER_ID = 1

def get_current_user():
    user = q1("SELECT * FROM users WHERE id=?", (DEFAULT_USER_ID,))
    if not user:
        user = q1("SELECT * FROM users ORDER BY id LIMIT 1")
    return user

def login_required(f):
    return f

def admin_required(f):
    return f

@app.route("/api/auth/me", methods=["GET", "OPTIONS"])
def me():
    if request.method == "OPTIONS":
        return "", 204
    if not session.get("csrf_token"):
        session["csrf_token"] = secrets.token_urlsafe(32)
        session.permanent = True
    csrf_token = session["csrf_token"]
    user = get_current_user()
    if not user:
        return jsonify({"authenticated": True, "id": 0, "name": "Guest", "initials": "?", "role": "admin", "color": "#4f8ef7", "csrf_token": csrf_token})
    return jsonify({
        "authenticated": True,
        "id": user["id"],
        "name": user["name"],
        "initials": user["initials"],
        "role": user["role"],
        "color": user["color"],
        "csrf_token": csrf_token,
    })

# ── QuickBooks ────────────────────────────────────────────────────────────────

QB_CLIENT_ID     = os.getenv("QB_CLIENT_ID", "")
QB_CLIENT_SECRET = os.getenv("QB_CLIENT_SECRET", "")
QB_REDIRECT_URI  = os.getenv("QB_REDIRECT_URI", "http://127.0.0.1:5000/qb/callback")
QB_REALM_ID      = os.getenv("QB_REALM_ID", "")
QB_ENVIRONMENT   = os.getenv("QB_ENVIRONMENT", "sandbox")
QB_TOKEN_URL     = "https://oauth.platform.intuit.com/oauth2/v1/tokens/bearer"
QB_AUTH_URL      = "https://appcenter.intuit.com/connect/oauth2"
QB_API_BASE      = (
    "https://sandbox-quickbooks.api.intuit.com/v3/company"
    if QB_ENVIRONMENT == "sandbox"
    else "https://quickbooks.api.intuit.com/v3/company"
)

def _ensure_qb_tokens_table():
    db = get_db()
    db.execute("CREATE TABLE IF NOT EXISTS qb_tokens (id INTEGER PRIMARY KEY AUTOINCREMENT, access_token TEXT, refresh_token TEXT, expires_at REAL, realm_id TEXT)")
    cols = {r[1] for r in db.execute("PRAGMA table_info(qb_tokens)").fetchall()}
    if "realm_id" not in cols:
        db.execute("ALTER TABLE qb_tokens ADD COLUMN realm_id TEXT")
    db.commit()

# OAuth state is also persisted here so the callback can validate even if the
# session cookie didn't survive the Intuit round-trip (e.g. user hit the app
# on localhost but the redirect_uri is 127.0.0.1, or Edge stripped SameSite=Lax
# on some flows). Rows expire after _QB_STATE_TTL seconds.
_QB_STATE_TTL = 900  # 15 minutes

def _ensure_qb_oauth_states_table():
    db = get_db()
    db.execute("CREATE TABLE IF NOT EXISTS qb_oauth_states (state TEXT PRIMARY KEY, created_at REAL)")
    db.commit()

def _save_oauth_state(state):
    _ensure_qb_oauth_states_table()
    db = get_db()
    now = datetime.now(timezone.utc).timestamp()
    db.execute("DELETE FROM qb_oauth_states WHERE created_at < ?", (now - _QB_STATE_TTL,))
    db.execute("INSERT OR REPLACE INTO qb_oauth_states (state, created_at) VALUES (?, ?)", (state, now))
    db.commit()

def _consume_oauth_state(state):
    _ensure_qb_oauth_states_table()
    db = get_db()
    now = datetime.now(timezone.utc).timestamp()
    row = db.execute("SELECT created_at FROM qb_oauth_states WHERE state=?", (state,)).fetchone()
    db.execute("DELETE FROM qb_oauth_states WHERE state=?", (state,))
    db.commit()
    if not row:
        return False
    return (now - row["created_at"]) <= _QB_STATE_TTL

def get_tokens():
    _ensure_qb_tokens_table()
    row = q1("SELECT access_token, refresh_token, expires_at, realm_id FROM qb_tokens ORDER BY id DESC LIMIT 1")
    if not row:
        return {}
    return {
        "access_token": decrypt_token(row["access_token"]),
        "refresh_token": decrypt_token(row["refresh_token"]),
        "expires_at": row["expires_at"],
        "realm_id": row["realm_id"],
    }

def save_tokens(at, rt, ei, realm_id=None):
    _ensure_qb_tokens_table()
    ea = datetime.now(timezone.utc).timestamp() + ei
    db = get_db()
    if realm_id is None:
        prev = q1("SELECT realm_id FROM qb_tokens ORDER BY id DESC LIMIT 1")
        realm_id = prev["realm_id"] if prev else None
    db.execute(
        "INSERT INTO qb_tokens (access_token, refresh_token, expires_at, realm_id) VALUES (?,?,?,?)",
        (encrypt_token(at), encrypt_token(rt), ea, realm_id),
    )
    db.commit()

def get_realm_id():
    tokens = get_tokens()
    return tokens.get("realm_id") or os.getenv("QB_REALM_ID", "")

def refresh_access_token():
    tokens = get_tokens()
    if not tokens.get("refresh_token"):
        return None
    resp = requests.post(QB_TOKEN_URL,
        data={"grant_type": "refresh_token", "refresh_token": tokens["refresh_token"]},
        auth=(QB_CLIENT_ID, QB_CLIENT_SECRET))
    if resp.ok:
        d = resp.json()
        save_tokens(d["access_token"], d.get("refresh_token", tokens["refresh_token"]), d["expires_in"])
        return d["access_token"]
    return None

def qb_get(path):
    tokens = get_tokens()
    if not tokens:
        return None, "Not connected"
    realm = get_realm_id()
    if not realm:
        return None, "No realm_id — reconnect to QuickBooks"
    now = datetime.now(timezone.utc).timestamp()
    token = tokens["access_token"]
    if tokens.get("expires_at", 0) < now + 60:
        token = refresh_access_token()
    if not token:
        return None, "Token refresh failed"
    resp = requests.get(
        f"{QB_API_BASE}/{realm}{path}",
        headers={"Authorization": f"Bearer {token}", "Accept": "application/json"})
    return (resp.json(), None) if resp.ok else (None, f"QB error {resp.status_code}: {resp.text[:200]}")

def _qb_error(reason):
    print(f"[qb] connect failed: {reason}", flush=True)
    return redirect("/?qb=error&reason=" + urllib.parse.quote(reason, safe=""))

@app.route("/api/qb/connect")
@login_required
def qb_connect():
    if not QB_CLIENT_ID or not QB_CLIENT_SECRET:
        return _qb_error("missing_credentials")
    state = secrets.token_urlsafe(16)
    session["qb_state"] = state
    _save_oauth_state(state)
    params = {
        "client_id": QB_CLIENT_ID,
        "scope": "com.intuit.quickbooks.accounting",
        "redirect_uri": QB_REDIRECT_URI,
        "response_type": "code",
        "state": state,
    }
    return redirect(QB_AUTH_URL + "?" + urllib.parse.urlencode(params))

@app.route("/qb/callback")
def qb_callback():
    code = request.args.get("code")
    realm = request.args.get("realmId")
    state = request.args.get("state")
    qb_err = request.args.get("error")
    if qb_err:
        return _qb_error(f"intuit_{qb_err}")
    if not code:
        return _qb_error("missing_code")
    if not state:
        return _qb_error("missing_state")
    session_state = session.get("qb_state")
    state_ok = (state == session_state) or _consume_oauth_state(state)
    if not state_ok:
        return _qb_error("state_mismatch")
    if not QB_CLIENT_ID or not QB_CLIENT_SECRET:
        return _qb_error("missing_credentials")
    try:
        resp = requests.post(QB_TOKEN_URL,
            data={"grant_type": "authorization_code", "code": code, "redirect_uri": QB_REDIRECT_URI},
            auth=(QB_CLIENT_ID, QB_CLIENT_SECRET),
            headers={"Accept": "application/json"},
            timeout=15)
    except requests.RequestException as e:
        return _qb_error(f"network:{type(e).__name__}")
    if not resp.ok:
        print(f"[qb] token exchange {resp.status_code}: {resp.text[:300]}", flush=True)
        return _qb_error(f"token_exchange_{resp.status_code}")
    try:
        d = resp.json()
    except ValueError:
        return _qb_error("token_response_not_json")
    if not d.get("access_token") or not d.get("refresh_token"):
        return _qb_error("token_response_incomplete")
    save_tokens(d["access_token"], d["refresh_token"], d.get("expires_in", 3600), realm_id=realm)
    session.pop("qb_state", None)
    return redirect("/?qb=connected")

@app.route("/api/qb/status", methods=["GET", "OPTIONS"])
@login_required
def qb_status():
    if request.method == "OPTIONS":
        return "", 204
    tokens = get_tokens()
    if not tokens:
        return jsonify({"connected": False})
    return jsonify({
        "connected": True,
        "token_expires_in": max(0, int(tokens.get("expires_at", 0) - datetime.now(timezone.utc).timestamp()))
    })

@app.route("/api/qb/diagnose", methods=["GET", "OPTIONS"])
@login_required
def qb_diagnose():
    """Fetches a report straight from QuickBooks and returns the raw shape so
    we can see *why* a sync produces zero lines. Usage:
    /api/qb/diagnose?period_id=123&rtype=pl   (rtype: pl | bs | cf; default pl)
    Returns: {connected, realm_id, environment, period, url, qb_error, rows,
              row_count_top_level, row_count_flattened, sample_rows, header}.
    No side effects — does not touch qb_reports / qb_report_lines."""
    if request.method == "OPTIONS":
        return "", 204
    tokens = get_tokens()
    realm = get_realm_id()
    out = {
        "connected": bool(tokens),
        "realm_id": realm or None,
        "environment": QB_ENVIRONMENT,
    }
    if not tokens:
        out["qb_error"] = "Not connected — click Settings → Connect QuickBooks"
        return jsonify(out)
    period_id = request.args.get("period_id", type=int)
    if not period_id:
        active = q1("SELECT id FROM periods WHERE is_active=1")
        period_id = active["id"] if active else None
    if not period_id:
        out["qb_error"] = "No period_id supplied and no active period"
        return jsonify(out)
    period = q1("SELECT id, label, start_date, end_date FROM periods WHERE id=?", (period_id,))
    if not period:
        out["qb_error"] = "Period not found"
        return jsonify(out)
    out["period"] = dict(period)
    rtype = (request.args.get("rtype") or "pl").lower()
    endpoints = {"pl": "ProfitAndLoss", "bs": "BalanceSheet", "cf": "CashFlow"}
    if rtype not in endpoints:
        out["qb_error"] = f"Unknown rtype '{rtype}' (expected pl/bs/cf)"
        return jsonify(out)
    path = f"/reports/{endpoints[rtype]}?start_date={period['start_date']}&end_date={period['end_date']}&accounting_method=Accrual&minorversion=65"
    out["url"] = f"{QB_API_BASE}/{realm}{path}"
    data, err_str = qb_get(path)
    if err_str:
        out["qb_error"] = err_str
        return jsonify(out)
    out["qb_error"] = None
    out["header"] = data.get("Header")
    rows = data.get("Rows", {}) or {}
    top_rows = rows.get("Row", []) if isinstance(rows, dict) else []
    out["row_count_top_level"] = len(top_rows)
    flat = []
    _flatten_qb_rows(rows, rtype, flat)
    out["row_count_flattened"] = len(flat)
    out["sample_rows"] = [
        {"name": r.get("account_name"), "amount": r.get("amount"),
         "section": r.get("section"), "is_subtotal": r.get("is_subtotal")}
        for r in flat[:12]
    ]
    return jsonify(out)

def sync_qb_accounts(period_id):
    """Pull QB's chart of accounts, cache it, and make sure there's a
    reconciliation row for every Asset/Liability/Equity account in the given
    period. Updates qb_balance on existing rows too. Idempotent: matches
    existing rows by AcctNum → Name → falls back to creating with QB Id as
    account_number."""
    _ensure_report_tables()
    data, error = qb_get("/query?query=SELECT%20*%20FROM%20Account%20MAXRESULTS%201000")
    if error:
        return {"ok": False, "error": error}
    accounts = data.get("QueryResponse", {}).get("Account", []) or []
    db = get_db()
    for a in accounts:
        db.execute("""INSERT INTO qb_accounts (id, name, acct_num, account_type, classification, synced_at)
                      VALUES (?,?,?,?,?,CURRENT_TIMESTAMP)
                      ON CONFLICT(id) DO UPDATE SET
                        name=excluded.name, acct_num=excluded.acct_num,
                        account_type=excluded.account_type, classification=excluded.classification,
                        synced_at=CURRENT_TIMESTAMP""",
                   (str(a.get("Id", "")), a.get("Name", ""), (a.get("AcctNum") or "").strip(),
                    a.get("AccountType", ""), a.get("Classification", "")))

    bs_accounts = [a for a in accounts if a.get("Classification") in ("Asset", "Liability", "Equity")]
    existing = q("SELECT id, account_number, account_name FROM reconciliations WHERE period_id=?", (period_id,))
    by_num  = {r["account_number"]: r["id"] for r in existing if (r["account_number"] or "").strip()}
    by_name = {r["account_name"].strip().lower(): r["id"] for r in existing if r["account_name"]}

    admin = q1("SELECT id FROM users WHERE role='admin' ORDER BY id LIMIT 1") \
            or q1("SELECT id FROM users ORDER BY id LIMIT 1")
    assignee_id = admin["id"] if admin else DEFAULT_USER_ID

    created = updated = 0
    for a in bs_accounts:
        acct_num = (a.get("AcctNum") or "").strip()
        name = (a.get("Name") or "").strip()
        if not name:
            continue
        try:
            bal = float(a.get("CurrentBalance") or 0)
        except (TypeError, ValueError):
            bal = 0.0
        rid = None
        if acct_num and acct_num in by_num:
            rid = by_num[acct_num]
        elif name.lower() in by_name:
            rid = by_name[name.lower()]
        if rid:
            db.execute("UPDATE reconciliations SET qb_balance=?, last_synced_at=CURRENT_TIMESTAMP WHERE id=?", (bal, rid))
            updated += 1
        else:
            db.execute("""INSERT INTO reconciliations
                          (period_id, account_number, account_name, assignee_id,
                           qb_balance, expected_balance, status, last_synced_at)
                          VALUES (?,?,?,?,?,NULL,'open',CURRENT_TIMESTAMP)""",
                       (period_id, acct_num or f"QB{a.get('Id','')}", name, assignee_id, bal))
            created += 1
    db.commit()
    return {"ok": True, "created": created, "updated": updated, "total": len(bs_accounts)}

def _ensure_period_close_columns():
    """Idempotent: add is_closed / closed_at / closed_by columns to `periods`
    if they're missing. Safe to call repeatedly."""
    db = get_db()
    cols = {r[1] for r in db.execute("PRAGMA table_info(periods)").fetchall()}
    for col, ddl in [
        ("is_closed", "ALTER TABLE periods ADD COLUMN is_closed INTEGER NOT NULL DEFAULT 0"),
        ("closed_at", "ALTER TABLE periods ADD COLUMN closed_at DATETIME"),
        ("closed_by", "ALTER TABLE periods ADD COLUMN closed_by INTEGER REFERENCES users(id)"),
    ]:
        if col not in cols:
            db.execute(ddl)
    db.commit()

def _backfill_closed_once():
    """On first-ever run (no period has been marked closed yet), mark every
    month period that ENDED BEFORE the most-recently-ended month as closed.
    Result: the earliest unclosed month is the most recently completed month.
    Joe's expected starting point — 'April 2026 today → start on March 2026'."""
    _ensure_period_close_columns()
    if q1("SELECT 1 FROM periods WHERE is_closed=1 LIMIT 1"):
        return  # user has already closed something; don't touch history
    today_iso = date.today().isoformat()
    latest_ended = q1("""SELECT end_date FROM periods
                         WHERE period_type='month' AND end_date < ?
                         ORDER BY end_date DESC LIMIT 1""", (today_iso,))
    if not latest_ended:
        return
    run("""UPDATE periods SET is_closed=1, closed_at=CURRENT_TIMESTAMP
           WHERE period_type='month' AND end_date < ?""", (latest_ended["end_date"],))

def _next_open_month():
    """The earliest month period with is_closed=0, globally. That's the close
    period the app should default to: 'whichever period has not been closed'.
    Returns a full row or None."""
    return q1("""SELECT * FROM periods
                 WHERE period_type='month' AND is_closed=0
                 ORDER BY start_date ASC LIMIT 1""")

def _activate_current_close_period():
    """Make the 'next open month' the active close period. Safe to call
    repeatedly; no-op if the active period is already the next-open-month.
    Returns the active period id (or None if no months exist)."""
    _ensure_fiscal_calendar()
    _ensure_period_close_columns()
    _backfill_closed_once()
    target = _next_open_month()
    if not target:
        active = q1("SELECT id FROM periods WHERE is_active=1")
        return active["id"] if active else None
    current = q1("SELECT id FROM periods WHERE is_active=1")
    if current and current["id"] == target["id"]:
        return target["id"]
    run("UPDATE periods SET is_active=0")
    run("UPDATE periods SET is_active=1 WHERE id=?", (target["id"],))
    return target["id"]

# Back-compat alias — older call sites still reference the old name.
_activate_current_period_if_stale = _activate_current_close_period

def sync_qb_balances():
    with app.app_context():
        period = q1("SELECT id FROM periods WHERE is_active=1")
        if not period:
            return {"ok": False, "error": "No active period"}
        return sync_qb_accounts(period["id"])

@app.route("/api/qb/sync", methods=["POST", "OPTIONS"])
@login_required
@csrf_protect
def manual_sync():
    if request.method == "OPTIONS":
        return "", 204
    return jsonify(sync_qb_balances())

@app.route("/api/qb/bootstrap", methods=["POST", "OPTIONS"])
@login_required
@csrf_protect
def qb_bootstrap():
    """One-click 'set this up from my real QuickBooks': activate the current
    4-4-5 month, seed reconciliations from QB's chart of accounts, pull
    P&L/BS/CF for current month/quarter/year, and then sweep every
    month/quarter/year from 2024-01-01 through today so historical reporting
    and prior-period reconciliations work. Body may pass
    {history_start: 'YYYY-MM-DD'} or {skip_history: true}."""
    if request.method == "OPTIONS":
        return "", 204
    if not get_tokens():
        return jsonify({"ok": False, "error": "Not connected to QuickBooks"}), 400
    period_id = _activate_current_period_if_stale()
    if not period_id:
        return jsonify({"ok": False, "error": "Could not activate a current period"}), 500
    accounts_result = sync_qb_accounts(period_id)
    active = q1("SELECT id, parent_id, fiscal_year FROM periods WHERE id=?", (period_id,))
    targets = [active["id"]]
    if active["parent_id"]:
        targets.append(active["parent_id"])
    year_row = q1("SELECT id FROM periods WHERE period_type='year' AND fiscal_year=?",
                  (active["fiscal_year"],))
    if year_row:
        targets.append(year_row["id"])
    report_results = {}
    for pid in targets:
        for rtype in ("pl", "bs", "cf"):
            report_results[f"{pid}:{rtype}"] = sync_qb_report(pid, rtype)

    body = request.get_json(silent=True) or {}
    history = None
    if not body.get("skip_history"):
        history = sync_history_range(
            start_iso=body.get("history_start") or "2024-01-01",
            end_iso=body.get("history_end") or date.today().isoformat(),
        )

    return jsonify({
        "ok": True,
        "period_id": period_id,
        "accounts": accounts_result,
        "reports": report_results,
        "history": history,
    })

# ── QuickBooks Report Sync (P&L, Balance Sheet) ───────────────────────────────

def _ensure_report_tables():
    db = get_db()
    db.execute("""CREATE TABLE IF NOT EXISTS qb_reports (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        period_id INTEGER NOT NULL REFERENCES periods(id),
        report_type TEXT NOT NULL,
        start_date TEXT NOT NULL,
        end_date TEXT NOT NULL,
        raw_json TEXT NOT NULL,
        pulled_at DATETIME DEFAULT CURRENT_TIMESTAMP,
        UNIQUE(period_id, report_type)
    )""")
    db.execute("""CREATE TABLE IF NOT EXISTS qb_report_lines (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        period_id INTEGER NOT NULL REFERENCES periods(id),
        report_type TEXT NOT NULL,
        section TEXT,
        account_name TEXT NOT NULL,
        account_id TEXT,
        amount REAL NOT NULL DEFAULT 0,
        is_subtotal INTEGER DEFAULT 0,
        depth INTEGER DEFAULT 0,
        sort_order INTEGER DEFAULT 0
    )""")
    cols = {r[1] for r in db.execute("PRAGMA table_info(qb_report_lines)").fetchall()}
    if "account_id" not in cols:
        db.execute("ALTER TABLE qb_report_lines ADD COLUMN account_id TEXT")
    db.execute("CREATE INDEX IF NOT EXISTS idx_rl_period_type ON qb_report_lines(period_id, report_type)")

    db.execute("""CREATE TABLE IF NOT EXISTS flux_notes (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        period_id INTEGER NOT NULL REFERENCES periods(id),
        report_type TEXT NOT NULL,
        account_name TEXT NOT NULL,
        note TEXT NOT NULL DEFAULT '',
        author_id INTEGER REFERENCES users(id),
        updated_at DATETIME DEFAULT CURRENT_TIMESTAMP,
        UNIQUE(period_id, report_type, account_name)
    )""")

    db.execute("""CREATE TABLE IF NOT EXISTS report_groups (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        name TEXT NOT NULL,
        report_type TEXT NOT NULL,
        sort_order INTEGER NOT NULL DEFAULT 0,
        created_at DATETIME DEFAULT CURRENT_TIMESTAMP
    )""")
    db.execute("""CREATE TABLE IF NOT EXISTS report_group_map (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        group_id INTEGER NOT NULL REFERENCES report_groups(id) ON DELETE CASCADE,
        account_name TEXT NOT NULL,
        UNIQUE(group_id, account_name)
    )""")

    db.execute("""CREATE TABLE IF NOT EXISTS qb_accounts (
        id TEXT PRIMARY KEY,
        name TEXT NOT NULL,
        acct_num TEXT,
        account_type TEXT,
        classification TEXT,
        synced_at DATETIME DEFAULT CURRENT_TIMESTAMP
    )""")
    db.commit()

def _flatten_qb_rows(rows, report_type, out, section=None, depth=0):
    """Recursively walk a QB Report Rows tree and emit flat line records."""
    if not rows or not isinstance(rows, dict):
        return
    for row in rows.get("Row", []):
        rtype = row.get("type", "Data")
        if rtype == "Section":
            header = row.get("Header", {}).get("ColData", [])
            header_name = header[0].get("value", "") if header else ""
            new_section = section or header_name or None
            _flatten_qb_rows(row.get("Rows", {}), report_type, out, new_section, depth+1)
            summary = row.get("Summary", {}).get("ColData", [])
            if summary:
                name = summary[0].get("value", "")
                try:
                    amt = float(summary[-1].get("value", 0) or 0)
                except (TypeError, ValueError):
                    amt = 0.0
                out.append({
                    "section": new_section, "account_name": name, "account_id": None, "amount": amt,
                    "is_subtotal": 1, "depth": depth, "sort_order": len(out),
                })
        else:
            cd = row.get("ColData", [])
            if not cd:
                continue
            name = cd[0].get("value", "")
            acct_id = cd[0].get("id")
            try:
                amt = float(cd[-1].get("value", 0) or 0)
            except (TypeError, ValueError):
                amt = 0.0
            out.append({
                "section": section, "account_name": name, "account_id": acct_id, "amount": amt,
                "is_subtotal": 0, "depth": depth, "sort_order": len(out),
            })

def sync_qb_report(period_id, report_type):
    """report_type: 'pl' | 'bs'. Pulls from QB and caches into qb_reports + qb_report_lines."""
    _ensure_report_tables()
    period = q1("SELECT id, start_date, end_date FROM periods WHERE id=?", (period_id,))
    if not period:
        return {"ok": False, "error": "Period not found"}

    if report_type == "pl":
        path = f"/reports/ProfitAndLoss?start_date={period['start_date']}&end_date={period['end_date']}&accounting_method=Accrual&minorversion=65"
    elif report_type == "bs":
        path = f"/reports/BalanceSheet?start_date={period['start_date']}&end_date={period['end_date']}&accounting_method=Accrual&minorversion=65"
    elif report_type == "cf":
        path = f"/reports/CashFlow?start_date={period['start_date']}&end_date={period['end_date']}&accounting_method=Accrual&minorversion=65"
    else:
        return {"ok": False, "error": "Unknown report_type"}

    data, error = qb_get(path)
    if error:
        return {"ok": False, "error": error}

    import json
    db = get_db()
    db.execute("DELETE FROM qb_report_lines WHERE period_id=? AND report_type=?", (period_id, report_type))
    db.execute("""INSERT INTO qb_reports (period_id, report_type, start_date, end_date, raw_json, pulled_at)
                  VALUES (?,?,?,?,?,CURRENT_TIMESTAMP)
                  ON CONFLICT(period_id, report_type) DO UPDATE SET
                    start_date=excluded.start_date, end_date=excluded.end_date,
                    raw_json=excluded.raw_json, pulled_at=CURRENT_TIMESTAMP""",
               (period_id, report_type, period["start_date"], period["end_date"], json.dumps(data)))

    lines = []
    _flatten_qb_rows(data.get("Rows", {}), report_type, lines)
    for ln in lines:
        db.execute("""INSERT INTO qb_report_lines
                      (period_id, report_type, section, account_name, account_id, amount, is_subtotal, depth, sort_order)
                      VALUES (?,?,?,?,?,?,?,?,?)""",
                   (period_id, report_type, ln["section"], ln["account_name"], ln.get("account_id"),
                    ln["amount"], ln["is_subtotal"], ln["depth"], ln["sort_order"]))
    db.commit()
    return {"ok": True, "lines": len(lines)}

def sync_qb_recons_from_bs(period_id):
    """Derive reconciliation rows for `period_id` from its cached Balance
    Sheet lines. Unlike sync_qb_accounts() — which stamps TODAY'S balance
    against a period — this uses the BS as-of the period end, so historical
    periods get their real ending balances. Call sync_qb_report(period_id,'bs')
    first. Idempotent: matches existing rows by name (case-insensitive)."""
    lines = q("""SELECT account_name, account_id, amount FROM qb_report_lines
                 WHERE period_id=? AND report_type='bs' AND is_subtotal=0
                   AND account_name != ''""", (period_id,))
    if not lines:
        return {"ok": False, "error": "No BS lines cached — run sync_qb_report(period_id,'bs') first"}

    db = get_db()
    existing = q("SELECT id, account_name FROM reconciliations WHERE period_id=?", (period_id,))
    by_name = {(r["account_name"] or "").strip().lower(): r["id"] for r in existing if r["account_name"]}

    qb_acct_by_id = {r["id"]: dict(r) for r in q("SELECT id, name, acct_num FROM qb_accounts")}

    admin = q1("SELECT id FROM users WHERE role='admin' ORDER BY id LIMIT 1") \
            or q1("SELECT id FROM users ORDER BY id LIMIT 1")
    assignee_id = admin["id"] if admin else DEFAULT_USER_ID

    created = updated = 0
    for ln in lines:
        name = (ln["account_name"] or "").strip()
        if not name:
            continue
        try:
            amt = float(ln["amount"] or 0)
        except (TypeError, ValueError):
            amt = 0.0
        acct_num = ""
        acct_id = ln["account_id"]
        if acct_id and str(acct_id) in qb_acct_by_id:
            acct_num = (qb_acct_by_id[str(acct_id)].get("acct_num") or "").strip()
        rid = by_name.get(name.lower())
        if rid:
            db.execute("UPDATE reconciliations SET qb_balance=?, last_synced_at=CURRENT_TIMESTAMP WHERE id=?",
                       (amt, rid))
            updated += 1
        else:
            db.execute("""INSERT INTO reconciliations
                          (period_id, account_number, account_name, assignee_id,
                           qb_balance, expected_balance, status, last_synced_at)
                          VALUES (?,?,?,?,?,NULL,'open',CURRENT_TIMESTAMP)""",
                       (period_id, acct_num or f"QB{acct_id or ''}", name, assignee_id, amt))
            created += 1
    db.commit()
    return {"ok": True, "created": created, "updated": updated, "total": len(lines)}

def sync_history_range(start_iso="2024-01-01", end_iso=None, types=("pl", "bs", "cf"), derive_recons=True):
    """Sweep every 4-4-5 month / quarter / year period whose window overlaps
    [start_iso, end_iso] and sync P&L / BS / CF into the cache. If derive_recons
    is on, also upsert per-period reconciliations from the cached BS lines.
    Returns a summary dict; non-fatal QB errors are collected in 'errors'."""
    _ensure_fiscal_calendar()
    _ensure_report_tables()
    if not end_iso:
        end_iso = date.today().isoformat()
    periods = q("""SELECT id, label, period_type, start_date, end_date FROM periods
                   WHERE end_date >= ? AND start_date <= ?
                   ORDER BY period_type, start_date""", (start_iso, end_iso))
    summary = {
        "ok": True, "start_date": start_iso, "end_date": end_iso,
        "periods": len(periods), "pl": 0, "bs": 0, "cf": 0,
        "recons_created": 0, "recons_updated": 0, "errors": [],
    }
    for p in periods:
        pid = p["id"]
        bs_ok = False
        for rtype in types:
            res = sync_qb_report(pid, rtype)
            if res.get("ok"):
                summary[rtype] = summary.get(rtype, 0) + 1
                if rtype == "bs":
                    bs_ok = True
            else:
                summary["errors"].append(f"{p['label']} {rtype}: {res.get('error')}")
        if derive_recons and bs_ok:
            rec = sync_qb_recons_from_bs(pid)
            if rec.get("ok"):
                summary["recons_created"] += rec.get("created", 0)
                summary["recons_updated"] += rec.get("updated", 0)
    return summary

@app.route("/api/qb/sync_history", methods=["POST", "OPTIONS"])
@login_required
@csrf_protect
def qb_sync_history():
    """One-click historical sweep. Body: {start_date?, end_date?, types?,
    derive_recons?}. Defaults to 2024-01-01 through today, all three reports,
    and deriving recons from each period's BS."""
    if request.method == "OPTIONS":
        return "", 204
    if not get_tokens():
        return jsonify({"ok": False, "error": "Not connected to QuickBooks"}), 400
    body = request.get_json(silent=True) or {}
    start = body.get("start_date") or "2024-01-01"
    end = body.get("end_date") or date.today().isoformat()
    types = body.get("types") or ("pl", "bs", "cf")
    derive = body.get("derive_recons", True)
    return jsonify(sync_history_range(start, end, types=tuple(types), derive_recons=bool(derive)))

@app.route("/api/qb/sync_reports", methods=["POST", "OPTIONS"])
@login_required
@csrf_protect
def sync_reports_endpoint():
    if request.method == "OPTIONS":
        return "", 204
    body = request.get_json(silent=True) or {}
    period_id = body.get("period_id")
    if not period_id:
        active = q1("SELECT id FROM periods WHERE is_active=1")
        period_id = active["id"] if active else None
    if not period_id:
        return jsonify({"ok": False, "error": "No period selected and no active period"}), 400
    types = body.get("types") or ["pl", "bs", "cf"]
    results = {t: sync_qb_report(period_id, t) for t in types}
    recons = None
    if "bs" in types and results.get("bs", {}).get("ok") and body.get("derive_recons", True):
        recons = sync_qb_recons_from_bs(period_id)
    return jsonify({
        "ok": all(r.get("ok") for r in results.values()),
        "results": results,
        "recons": recons,
    })

def _load_report_lines(period_id, report_type):
    rows = q("""SELECT section, account_name, account_id, amount, is_subtotal, depth, sort_order
                FROM qb_report_lines WHERE period_id=? AND report_type=?
                ORDER BY sort_order""", (period_id, report_type))
    return [dict(r) for r in rows]

def _prior_period_id(period_id, mode="prev"):
    """Resolve a compare-to period id from a mode string. Supported modes:
      prev    — immediately prior period of same type (MoM / QoQ / Y-1)
      prev2   — two periods ago (same type)
      prev3   — three periods ago (same type)
      yoy     — same type + same period_number in the previous fiscal year
      yoy2    — same period_number two fiscal years ago
      ytd     — the FY `year` period of the same fiscal year (this period vs YTD)
      ytd_ly  — the FY `year` period of the prior fiscal year
    Unknown modes fall back to 'prev'."""
    cur = q1("SELECT start_date, period_type, period_number, fiscal_year FROM periods WHERE id=?", (period_id,))
    if not cur:
        return None
    ptype = cur["period_type"] or "month"
    fy = cur["fiscal_year"]

    if mode in ("yoy", "yoy2") and fy:
        back = 1 if mode == "yoy" else 2
        pn_match = "period_number IS NULL" if cur["period_number"] is None else "period_number=?"
        params = [ptype, fy - back]
        if cur["period_number"] is not None:
            params.append(cur["period_number"])
        row = q1(f"""SELECT id FROM periods
                     WHERE period_type=? AND fiscal_year=? AND {pn_match}
                     LIMIT 1""", tuple(params))
        if row:
            return row["id"]

    if mode == "ytd" and fy:
        row = q1("SELECT id FROM periods WHERE period_type='year' AND fiscal_year=? LIMIT 1", (fy,))
        if row:
            return row["id"]
    if mode == "ytd_ly" and fy:
        row = q1("SELECT id FROM periods WHERE period_type='year' AND fiscal_year=? LIMIT 1", (fy - 1,))
        if row:
            return row["id"]

    offset_map = {"prev": 1, "prev2": 2, "prev3": 3}
    offset = offset_map.get(mode, 1)
    # LIMIT 1 OFFSET N-1 picks the Nth-prior period of this type
    row = q1(f"""SELECT id FROM periods WHERE period_type=? AND start_date < ?
                 ORDER BY start_date DESC LIMIT 1 OFFSET {offset - 1}""",
             (ptype, cur["start_date"]))
    return row["id"] if row else None

def _build_report_payload(period_id, rtype, compare_id=None, view="native"):
    cur_lines = _load_report_lines(period_id, rtype)
    cmp_lines = _load_report_lines(compare_id, rtype) if compare_id else []
    cmp_by_name = {(l["section"], l["account_name"]): l["amount"] for l in cmp_lines}

    notes = {r["account_name"]: {"note": r["note"], "updated_at": r["updated_at"], "author_id": r["author_id"]}
             for r in q("""SELECT account_name, note, updated_at, author_id FROM flux_notes
                           WHERE period_id=? AND report_type=?""", (period_id, rtype))}

    merged = []
    for l in cur_lines:
        prior_amt = cmp_by_name.get((l["section"], l["account_name"]))
        var = None if prior_amt is None else (l["amount"] - prior_amt)
        var_pct = None
        if prior_amt not in (None, 0):
            var_pct = round((l["amount"] - prior_amt) / abs(prior_amt) * 100, 2)
        n = notes.get(l["account_name"], {})
        merged.append({**l, "prior_amount": prior_amt, "variance": var, "variance_pct": var_pct,
                       "flux_note": n.get("note", ""), "flux_updated_at": n.get("updated_at"),
                       "flux_author_id": n.get("author_id")})

    if view == "custom":
        merged = _regroup_lines_custom(period_id, rtype, merged, compare_id)

    return merged

def _regroup_lines_custom(period_id, rtype, lines, compare_id=None):
    """Re-group non-subtotal lines under admin-defined groups; preserve uncategorized under 'Other'."""
    groups = q("SELECT id, name, sort_order FROM report_groups WHERE report_type=? ORDER BY sort_order, name", (rtype,))
    if not groups:
        return lines
    name_to_group = {}
    for g in groups:
        for m in q("SELECT account_name FROM report_group_map WHERE group_id=?", (g["id"],)):
            name_to_group[m["account_name"]] = g["name"]

    buckets = {g["name"]: [] for g in groups}
    buckets["Other"] = []
    for l in lines:
        if l.get("is_subtotal"):
            continue
        g_name = name_to_group.get(l["account_name"], "Other")
        buckets.setdefault(g_name, []).append(l)

    out = []
    sort_order = 0
    group_order = [g["name"] for g in groups] + ["Other"]
    for gname in group_order:
        rows = buckets.get(gname, [])
        if not rows:
            continue
        for r in rows:
            out.append({**r, "section": gname, "depth": 1, "sort_order": sort_order, "is_subtotal": 0})
            sort_order += 1
        subtotal = sum(r["amount"] for r in rows)
        prior_subtotal = sum((r.get("prior_amount") or 0) for r in rows if r.get("prior_amount") is not None)
        has_prior = any(r.get("prior_amount") is not None for r in rows)
        var = (subtotal - prior_subtotal) if has_prior else None
        var_pct = None
        if has_prior and prior_subtotal != 0:
            var_pct = round((subtotal - prior_subtotal) / abs(prior_subtotal) * 100, 2)
        out.append({
            "section": gname, "account_name": f"Total {gname}", "account_id": None,
            "amount": subtotal, "is_subtotal": 1, "depth": 0, "sort_order": sort_order,
            "prior_amount": prior_subtotal if has_prior else None,
            "variance": var, "variance_pct": var_pct,
            "flux_note": "", "flux_updated_at": None, "flux_author_id": None,
        })
        sort_order += 1
    return out

@app.route("/api/reports/<rtype>", methods=["GET", "OPTIONS"])
@login_required
def get_report(rtype):
    if request.method == "OPTIONS":
        return "", 204
    if rtype not in ("pl", "bs", "cf"):
        return jsonify({"error": "Unknown report type"}), 400
    _ensure_report_tables()
    period_id = request.args.get("period_id", type=int)
    if not period_id:
        active = q1("SELECT id FROM periods WHERE is_active=1")
        period_id = active["id"] if active else None
    if not period_id:
        return jsonify({"error": "No period"}), 400

    compare_id = request.args.get("compare_to", type=int)
    if compare_id is None and request.args.get("auto_compare", "1") == "1":
        compare_id = _prior_period_id(period_id, request.args.get("compare_mode", "prev"))

    view = request.args.get("view", "native")
    merged = _build_report_payload(period_id, rtype, compare_id, view)

    period_meta = q1("SELECT id, label, start_date, end_date FROM periods WHERE id=?", (period_id,))
    compare_meta = q1("SELECT id, label, start_date, end_date FROM periods WHERE id=?", (compare_id,)) if compare_id else None
    pulled = q1("SELECT pulled_at FROM qb_reports WHERE period_id=? AND report_type=?", (period_id, rtype))

    return jsonify({
        "report_type": rtype,
        "view": view,
        "period": dict(period_meta) if period_meta else None,
        "compare": dict(compare_meta) if compare_meta else None,
        "pulled_at": pulled["pulled_at"] if pulled else None,
        "lines": merged,
    })

# ── Transaction drill-down ────────────────────────────────────────────────────

@app.route("/api/qb/transactions", methods=["GET", "OPTIONS"])
@login_required
def qb_transactions():
    if request.method == "OPTIONS":
        return "", 204
    account_id = request.args.get("account_id")
    account_name = request.args.get("account_name")
    period_id = request.args.get("period_id", type=int)
    if not period_id:
        return jsonify({"error": "period_id required"}), 400
    period = q1("SELECT start_date, end_date FROM periods WHERE id=?", (period_id,))
    if not period:
        return jsonify({"error": "Period not found"}), 404

    if not account_id and account_name:
        row = q1("""SELECT account_id FROM qb_report_lines
                    WHERE period_id=? AND account_name=? AND account_id IS NOT NULL
                    LIMIT 1""", (period_id, account_name))
        if row:
            account_id = row["account_id"]

    qs = [f"start_date={period['start_date']}", f"end_date={period['end_date']}",
          "accounting_method=Accrual", "minorversion=65"]
    if account_id:
        qs.append(f"account={account_id}")
    data, error = qb_get("/reports/TransactionList?" + "&".join(qs))
    if error:
        return jsonify({"error": error}), 502

    cols = data.get("Columns", {}).get("Column", [])
    col_titles = [c.get("ColTitle", "") for c in cols]
    flat = []
    def _walk(rows):
        if not isinstance(rows, dict):
            return
        for row in rows.get("Row", []):
            if row.get("type") == "Section":
                _walk(row.get("Rows", {}))
            else:
                cd = row.get("ColData", [])
                flat.append({col_titles[i] if i < len(col_titles) else f"col{i}": (cd[i].get("value") if i < len(cd) else "")
                             for i in range(max(len(cd), len(col_titles)))})
    _walk(data.get("Rows", {}))

    return jsonify({
        "account_id": account_id,
        "account_name": account_name,
        "period_id": period_id,
        "columns": col_titles,
        "transactions": flat,
    })

# ── Flux Notes ────────────────────────────────────────────────────────────────

@app.route("/api/flux_notes", methods=["GET", "POST", "OPTIONS"])
@login_required
@csrf_protect
def flux_notes():
    if request.method == "OPTIONS":
        return "", 204
    _ensure_report_tables()
    if request.method == "GET":
        period_id = request.args.get("period_id", type=int)
        rtype = request.args.get("report_type")
        if not period_id or not rtype:
            return jsonify({"error": "period_id and report_type required"}), 400
        rows = q("""SELECT id, account_name, note, author_id, updated_at FROM flux_notes
                    WHERE period_id=? AND report_type=?""", (period_id, rtype))
        return jsonify([dict(r) for r in rows])

    body = request.get_json(silent=True) or {}
    period_id = body.get("period_id")
    rtype = body.get("report_type")
    account_name = body.get("account_name")
    note = (body.get("note") or "").strip()
    if not (period_id and rtype and account_name):
        return jsonify({"error": "period_id, report_type, account_name required"}), 400
    user = get_current_user()
    uid = user["id"] if user else DEFAULT_USER_ID
    if not note:
        run("DELETE FROM flux_notes WHERE period_id=? AND report_type=? AND account_name=?",
            (period_id, rtype, account_name))
        return jsonify({"ok": True, "deleted": True})
    run("""INSERT INTO flux_notes (period_id, report_type, account_name, note, author_id, updated_at)
           VALUES (?,?,?,?,?,CURRENT_TIMESTAMP)
           ON CONFLICT(period_id, report_type, account_name) DO UPDATE SET
             note=excluded.note, author_id=excluded.author_id, updated_at=CURRENT_TIMESTAMP""",
        (period_id, rtype, account_name, note, uid))
    return jsonify({"ok": True})

# ── Report Groups (custom P&L / BS grouping) ──────────────────────────────────

@app.route("/api/report_groups", methods=["GET", "POST", "OPTIONS"])
@login_required
@csrf_protect
def report_groups():
    if request.method == "OPTIONS":
        return "", 204
    _ensure_report_tables()
    if request.method == "GET":
        rtype = request.args.get("report_type")
        sql = "SELECT id, name, report_type, sort_order FROM report_groups"
        params = ()
        if rtype:
            sql += " WHERE report_type=?"
            params = (rtype,)
        sql += " ORDER BY report_type, sort_order, name"
        groups = [dict(r) for r in q(sql, params)]
        for g in groups:
            g["accounts"] = [r["account_name"] for r in q(
                "SELECT account_name FROM report_group_map WHERE group_id=? ORDER BY account_name", (g["id"],))]
        return jsonify(groups)

    u = get_current_user()
    if not u or u["role"] != "admin":
        return jsonify({"error": "Admin access required"}), 403
    body = request.get_json(silent=True) or {}
    name = (body.get("name") or "").strip()
    rtype = body.get("report_type")
    if not name or rtype not in ("pl", "bs", "cf"):
        return jsonify({"error": "name and valid report_type required"}), 400
    cur = run("INSERT INTO report_groups (name, report_type, sort_order) VALUES (?,?,?)",
              (name, rtype, int(body.get("sort_order") or 0)))
    return jsonify({"id": cur.lastrowid, "name": name, "report_type": rtype})

@app.route("/api/report_groups/<int:gid>", methods=["PATCH", "DELETE", "OPTIONS"])
@login_required
@csrf_protect
def report_group_detail(gid):
    if request.method == "OPTIONS":
        return "", 204
    u = get_current_user()
    if not u or u["role"] != "admin":
        return jsonify({"error": "Admin access required"}), 403
    _ensure_report_tables()
    if request.method == "DELETE":
        run("DELETE FROM report_group_map WHERE group_id=?", (gid,))
        run("DELETE FROM report_groups WHERE id=?", (gid,))
        return jsonify({"ok": True})
    body = request.get_json(silent=True) or {}
    if "name" in body:
        run("UPDATE report_groups SET name=? WHERE id=?", (body["name"], gid))
    if "sort_order" in body:
        run("UPDATE report_groups SET sort_order=? WHERE id=?", (int(body["sort_order"]), gid))
    if "accounts" in body:
        run("DELETE FROM report_group_map WHERE group_id=?", (gid,))
        for acct in body["accounts"]:
            run("INSERT OR IGNORE INTO report_group_map (group_id, account_name) VALUES (?,?)", (gid, acct))
    return jsonify({"ok": True})

# ── KPI Dashboard ─────────────────────────────────────────────────────────────

def _find_line(lines, *needles):
    needles = [n.lower() for n in needles]
    for l in lines:
        nm = (l.get("account_name") or "").lower().strip()
        if any(n == nm or nm.endswith(n) or nm.startswith(n) for n in needles):
            return l["amount"]
    return None

@app.route("/api/reports/kpis", methods=["GET", "OPTIONS"])
@login_required
def report_kpis():
    if request.method == "OPTIONS":
        return "", 204
    _ensure_report_tables()
    period_id = request.args.get("period_id", type=int)
    if not period_id:
        active = q1("SELECT id FROM periods WHERE is_active=1")
        period_id = active["id"] if active else None
    if not period_id:
        return jsonify({"error": "No period"}), 400

    pl = _load_report_lines(period_id, "pl")
    bs = _load_report_lines(period_id, "bs")

    revenue = _find_line(pl, "total income", "total revenue")
    cogs = _find_line(pl, "total cost of goods sold", "total cogs")
    gross = _find_line(pl, "gross profit")
    opex = _find_line(pl, "total expenses", "total operating expenses")
    net_income = _find_line(pl, "net income")

    total_assets = _find_line(bs, "total assets")
    current_assets = _find_line(bs, "total current assets")
    current_liab = _find_line(bs, "total current liabilities")
    total_liab = _find_line(bs, "total liabilities")
    equity = _find_line(bs, "total equity", "total stockholders' equity")
    cash = _find_line(bs, "total bank accounts", "total cash and cash equivalents")
    inventory = _find_line(bs, "total other current assets", "total inventory")

    def div(a, b):
        return round(a / b, 4) if (a is not None and b not in (None, 0)) else None

    kpis = {
        "revenue": revenue,
        "cogs": cogs,
        "gross_profit": gross,
        "operating_expenses": opex,
        "net_income": net_income,
        "gross_margin_pct": round(div(gross, revenue) * 100, 2) if div(gross, revenue) is not None else None,
        "operating_margin_pct": round(div((gross or 0) - (opex or 0), revenue) * 100, 2) if revenue else None,
        "net_margin_pct": round(div(net_income, revenue) * 100, 2) if div(net_income, revenue) is not None else None,
        "total_assets": total_assets,
        "total_liabilities": total_liab,
        "total_equity": equity,
        "cash": cash,
        "current_ratio": div(current_assets, current_liab),
        "quick_ratio": div((current_assets or 0) - (inventory or 0), current_liab) if current_liab else None,
        "debt_to_equity": div(total_liab, equity),
        "working_capital": (current_assets - current_liab) if (current_assets is not None and current_liab is not None) else None,
    }
    return jsonify({"period_id": period_id, "kpis": kpis})

# ── Report Export ─────────────────────────────────────────────────────────────

@app.route("/api/reports/<rtype>/export", methods=["GET"])
@login_required
def export_report(rtype):
    if rtype not in ("pl", "bs", "cf"):
        return "Unknown report type", 400
    _ensure_report_tables()
    period_id = request.args.get("period_id", type=int)
    if not period_id:
        active = q1("SELECT id FROM periods WHERE is_active=1")
        period_id = active["id"] if active else None
    if not period_id:
        return "No period", 400
    compare_id = request.args.get("compare_to", type=int)
    if compare_id is None and request.args.get("auto_compare", "1") == "1":
        compare_id = _prior_period_id(period_id, request.args.get("compare_mode", "prev"))
    view = request.args.get("view", "native")
    fmt = (request.args.get("format") or "xlsx").lower()

    lines = _build_report_payload(period_id, rtype, compare_id, view)
    period = q1("SELECT label FROM periods WHERE id=?", (period_id,))
    compare = q1("SELECT label FROM periods WHERE id=?", (compare_id,)) if compare_id else None
    title = {"pl": "Profit & Loss", "bs": "Balance Sheet", "cf": "Cash Flow"}[rtype]
    headers = ["Account", f"{period['label'] if period else 'Current'}"]
    if compare:
        headers += [compare["label"], "Variance $", "Variance %"]
    headers.append("Flux Note")

    def row_for(l):
        r = [("  " * (l.get("depth") or 0)) + (l.get("account_name") or ""), l.get("amount")]
        if compare:
            r += [l.get("prior_amount"), l.get("variance"), l.get("variance_pct")]
        r.append(l.get("flux_note") or "")
        return r

    if fmt == "csv":
        import csv, io
        buf = io.StringIO()
        w = csv.writer(buf)
        w.writerow([title, period["label"] if period else ""])
        w.writerow([])
        w.writerow(headers)
        for l in lines:
            w.writerow(row_for(l))
        from flask import Response
        return Response(buf.getvalue(), mimetype="text/csv",
                        headers={"Content-Disposition": f'attachment; filename="{rtype}_{period_id}.csv"'})

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        return "openpyxl not installed — run: pip install openpyxl  (or use ?format=csv)", 500

    wb = Workbook()
    ws = wb.active
    ws.title = title[:31]
    bold = Font(bold=True)
    hdr_fill = PatternFill("solid", fgColor="1E2D3D")
    hdr_font = Font(bold=True, color="E2E8F0")
    sub_fill = PatternFill("solid", fgColor="F1F5F9")
    right = Alignment(horizontal="right")
    thin = Side(border_style="thin", color="CBD5E1")

    ws.cell(1, 1, title).font = Font(bold=True, size=14)
    ws.cell(2, 1, f"Period: {period['label'] if period else ''}").font = Font(italic=True, color="64748B")
    if compare:
        ws.cell(2, 2, f"Compare: {compare['label']}").font = Font(italic=True, color="64748B")

    for i, h in enumerate(headers, 1):
        c = ws.cell(4, i, h)
        c.font = hdr_font
        c.fill = hdr_fill
        c.alignment = right if i > 1 else Alignment(horizontal="left")
        c.border = Border(bottom=thin)

    for ri, l in enumerate(lines, 5):
        row = row_for(l)
        for ci, val in enumerate(row, 1):
            c = ws.cell(ri, ci, val)
            if ci > 1 and ci < len(row):
                c.number_format = '#,##0.00;[Red](#,##0.00)' if ci != len(row)-1 or not compare else '0.00"%"'
                c.alignment = right
            if l.get("is_subtotal"):
                c.font = bold
                c.fill = sub_fill
        if compare and l.get("variance_pct") is not None:
            ws.cell(ri, len(headers)-1).number_format = '0.00"%"'

    ws.column_dimensions['A'].width = 42
    for col_letter in "BCDEF":
        ws.column_dimensions[col_letter].width = 18
    ws.column_dimensions['G' if compare else 'C'].width = 60

    import io
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    from flask import Response
    return Response(buf.read(),
                    mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    headers={"Content-Disposition": f'attachment; filename="{rtype}_{period_id}.xlsx"'})

# ── Fiscal Calendar (4-4-5) ───────────────────────────────────────────────────

from datetime import date, timedelta

_MONTH_NAMES = ["January", "February", "March", "April", "May", "June",
                "July", "August", "September", "October", "November", "December"]

def _iso_week1_monday(year):
    """Monday of ISO week 1 for given year (the week containing Jan 4)."""
    jan4 = date(year, 1, 4)
    return jan4 - timedelta(days=jan4.weekday())

def _generate_445_year(year):
    """Yields (period_type, period_number, label, start, end) for one 4-4-5 fiscal year."""
    w1_mon = _iso_week1_monday(year)
    pattern = [4, 4, 5, 4, 4, 5, 4, 4, 5, 4, 4, 5]
    out, month_ranges, week_idx = [], [], 0
    for i, weeks in enumerate(pattern):
        start = w1_mon + timedelta(days=week_idx * 7)
        end = start + timedelta(days=weeks * 7 - 1)
        mnum = i + 1
        if year == 2023 and mnum == 1:
            start = date(2023, 1, 1)
        month_ranges.append((start, end))
        out.append(("month", mnum, f"{_MONTH_NAMES[i]} {year}", start, end))
        week_idx += weeks
    for q in range(4):
        qstart = month_ranges[q * 3][0]
        qend = month_ranges[q * 3 + 2][1]
        out.append(("quarter", q + 1, f"Q{q+1} {year}", qstart, qend))
    out.append(("year", None, f"FY {year}", month_ranges[0][0], month_ranges[11][1]))
    return out

def _generate_gregorian_year(year):
    """Yields standard-calendar periods for one year (used pre-2023)."""
    out, month_ranges = [], []
    for m in range(1, 13):
        start = date(year, m, 1)
        end = (date(year, m + 1, 1) - timedelta(days=1)) if m < 12 else date(year, 12, 31)
        month_ranges.append((start, end))
        out.append(("month", m, f"{_MONTH_NAMES[m-1]} {year}", start, end))
    for q in range(4):
        out.append(("quarter", q + 1, f"Q{q+1} {year}",
                    month_ranges[q * 3][0], month_ranges[q * 3 + 2][1]))
    out.append(("year", None, f"FY {year}", date(year, 1, 1), date(year, 12, 31)))
    return out

def _ensure_fiscal_calendar():
    """Idempotent: adds fiscal columns to periods, seeds 2022-2030 hierarchy."""
    db = get_db()
    cols = {r[1] for r in db.execute("PRAGMA table_info(periods)").fetchall()}
    for col, ddl in [
        ("period_type", "ALTER TABLE periods ADD COLUMN period_type TEXT DEFAULT 'month'"),
        ("fiscal_year", "ALTER TABLE periods ADD COLUMN fiscal_year INTEGER"),
        ("period_number", "ALTER TABLE periods ADD COLUMN period_number INTEGER"),
        ("parent_id", "ALTER TABLE periods ADD COLUMN parent_id INTEGER"),
        ("calendar_type", "ALTER TABLE periods ADD COLUMN calendar_type TEXT DEFAULT 'gregorian'"),
    ]:
        if col not in cols:
            db.execute(ddl)
    db.commit()

    have_q = q1("SELECT COUNT(*) c FROM periods WHERE period_type='quarter'")
    if have_q and have_q["c"] > 0:
        return  # already seeded

    for year in range(2022, 2031):
        cal_type = "gregorian" if year < 2023 else "4-4-5"
        gen = _generate_gregorian_year if year < 2023 else _generate_445_year
        periods_for_year = gen(year)

        year_p = next(p for p in periods_for_year if p[0] == "year")
        quarter_ps = [p for p in periods_for_year if p[0] == "quarter"]
        month_ps = [p for p in periods_for_year if p[0] == "month"]

        cur = db.execute("""INSERT INTO periods
            (label, start_date, end_date, is_active, period_type, fiscal_year,
             period_number, parent_id, calendar_type)
            VALUES (?,?,?,0,?,?,?,?,?)""",
            (year_p[2], year_p[3].isoformat(), year_p[4].isoformat(),
             "year", year, None, None, cal_type))
        year_id = cur.lastrowid

        quarter_ids = {}
        for qp in quarter_ps:
            cur = db.execute("""INSERT INTO periods
                (label, start_date, end_date, is_active, period_type, fiscal_year,
                 period_number, parent_id, calendar_type)
                VALUES (?,?,?,0,?,?,?,?,?)""",
                (qp[2], qp[3].isoformat(), qp[4].isoformat(),
                 "quarter", year, qp[1], year_id, cal_type))
            quarter_ids[qp[1]] = cur.lastrowid

        for mp in month_ps:
            mnum = mp[1]
            qnum = (mnum - 1) // 3 + 1
            parent_q_id = quarter_ids[qnum]
            existing = q1("SELECT id FROM periods WHERE label=? AND period_type='month'", (mp[2],))
            if existing:
                db.execute("""UPDATE periods SET start_date=?, end_date=?,
                              period_type='month', fiscal_year=?, period_number=?,
                              parent_id=?, calendar_type=? WHERE id=?""",
                           (mp[3].isoformat(), mp[4].isoformat(), year, mnum,
                            parent_q_id, cal_type, existing["id"]))
            else:
                db.execute("""INSERT INTO periods
                    (label, start_date, end_date, is_active, period_type, fiscal_year,
                     period_number, parent_id, calendar_type)
                    VALUES (?,?,?,0,?,?,?,?,?)""",
                    (mp[2], mp[3].isoformat(), mp[4].isoformat(),
                     "month", year, mnum, parent_q_id, cal_type))
    db.commit()

# ── Auto-sync all reports every 15 min ────────────────────────────────────────

def sync_qb_all_reports():
    """Background: refresh P&L / BS / CF for active month, its quarter, and its year."""
    with app.app_context():
        if not get_tokens():
            return
        _ensure_fiscal_calendar()
        active = q1("SELECT id, parent_id, fiscal_year FROM periods WHERE is_active=1 AND period_type='month'")
        if not active:
            return
        targets = [active["id"]]
        if active["parent_id"]:
            targets.append(active["parent_id"])
        year_row = q1("SELECT id FROM periods WHERE period_type='year' AND fiscal_year=?",
                      (active["fiscal_year"],))
        if year_row:
            targets.append(year_row["id"])
        for pid in targets:
            for rtype in ("pl", "bs", "cf"):
                try:
                    sync_qb_report(pid, rtype)
                except Exception:
                    pass

@app.route("/api/calendar/reseed", methods=["POST", "OPTIONS"])
@login_required
@csrf_protect
def reseed_calendar():
    if request.method == "OPTIONS":
        return "", 204
    u = get_current_user()
    if not u or u["role"] != "admin":
        return jsonify({"error": "Admin only"}), 403
    _ensure_fiscal_calendar()
    return jsonify({"ok": True})

scheduler = BackgroundScheduler()
scheduler.add_job(sync_qb_balances, "interval", minutes=15, id="qb_sync")
scheduler.add_job(sync_qb_all_reports, "interval", minutes=15, id="qb_reports_sync")
scheduler.start()

def _ensure_new_features_schema():
    """Idempotent: add all columns/tables for the 13 new features."""
    db = get_db()

    # Feature 2: task dependencies
    db.execute("""CREATE TABLE IF NOT EXISTS task_dependencies (
        task_id INTEGER NOT NULL REFERENCES tasks(id) ON DELETE CASCADE,
        depends_on_id INTEGER NOT NULL REFERENCES tasks(id) ON DELETE CASCADE,
        PRIMARY KEY(task_id, depends_on_id)
    )""")

    # Feature 4: checklist templates
    db.execute("""CREATE TABLE IF NOT EXISTS checklist_templates (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        name TEXT NOT NULL,
        period_type TEXT DEFAULT 'monthly',
        description TEXT,
        created_by INTEGER REFERENCES users(id),
        created_at DATETIME DEFAULT CURRENT_TIMESTAMP
    )""")
    db.execute("""CREATE TABLE IF NOT EXISTS template_tasks (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        template_id INTEGER NOT NULL REFERENCES checklist_templates(id) ON DELETE CASCADE,
        title TEXT NOT NULL,
        category_id INTEGER REFERENCES categories(id),
        assignee_id INTEGER REFERENCES users(id),
        reviewer_id INTEGER REFERENCES users(id),
        day_target INTEGER,
        sort_order INTEGER DEFAULT 0
    )""")

    # Feature 6: open items
    db.execute("""CREATE TABLE IF NOT EXISTS open_items (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        period_id INTEGER NOT NULL REFERENCES periods(id),
        title TEXT NOT NULL,
        description TEXT,
        status TEXT DEFAULT 'open',
        priority TEXT DEFAULT 'medium',
        assigned_to INTEGER REFERENCES users(id),
        created_by INTEGER REFERENCES users(id),
        created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
        resolved_at DATETIME,
        resolved_by INTEGER REFERENCES users(id)
    )""")

    # Feature 8: attachments
    db.execute("""CREATE TABLE IF NOT EXISTS attachments (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        task_id INTEGER REFERENCES tasks(id) ON DELETE CASCADE,
        recon_id INTEGER REFERENCES reconciliations(id) ON DELETE CASCADE,
        filename TEXT NOT NULL,
        content_type TEXT,
        data BLOB NOT NULL,
        uploaded_by INTEGER REFERENCES users(id),
        uploaded_at DATETIME DEFAULT CURRENT_TIMESTAMP
    )""")

    # Feature 9: journal entries
    db.execute("""CREATE TABLE IF NOT EXISTS journal_entries (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        period_id INTEGER NOT NULL REFERENCES periods(id),
        je_number TEXT,
        description TEXT NOT NULL,
        debit_account TEXT,
        credit_account TEXT,
        amount REAL,
        preparer_id INTEGER REFERENCES users(id),
        reviewer_id INTEGER REFERENCES users(id),
        status TEXT DEFAULT 'draft',
        prepared_at DATETIME,
        submitted_at DATETIME,
        approved_at DATETIME,
        notes TEXT
    )""")

    # Feature 5: recurrence on tasks
    cols = {r[1] for r in db.execute("PRAGMA table_info(tasks)").fetchall()}
    if "recurrence" not in cols:
        db.execute("ALTER TABLE tasks ADD COLUMN recurrence TEXT DEFAULT 'none'")
    # Feature 7: submitted_at/submitted_by on tasks
    if "submitted_at" not in cols:
        db.execute("ALTER TABLE tasks ADD COLUMN submitted_at DATETIME")
    if "submitted_by" not in cols:
        db.execute("ALTER TABLE tasks ADD COLUMN submitted_by INTEGER REFERENCES users(id)")

    db.commit()


# Ensure the base schema + seed data exist before the fiscal-calendar migration
# runs. init_db.init() is idempotent.
try:
    from init_db import init as _init_db
    _init_db()
    with app.app_context():
        _ensure_fiscal_calendar()
        _ensure_period_close_columns()
        _backfill_closed_once()
        _activate_current_close_period()
        _ensure_new_features_schema()
except Exception:
    traceback.print_exc()

# ── Periods ───────────────────────────────────────────────────────────────────

@app.route("/api/periods", methods=["GET", "OPTIONS"])
@login_required
def get_periods():
    if request.method == "OPTIONS":
        return "", 204
    return jsonify(rows_to_list(q("SELECT * FROM periods ORDER BY start_date DESC")))

@app.route("/api/periods/active", methods=["GET", "OPTIONS"])
@login_required
def get_active_period():
    if request.method == "OPTIONS":
        return "", 204
    row = q1("SELECT * FROM periods WHERE is_active=1")
    return jsonify(dict(row)) if row else err("No active period", 404)

@app.route("/api/periods", methods=["POST", "OPTIONS"])
@admin_required
@csrf_protect
def create_period():
    if request.method == "OPTIONS":
        return "", 204
    b = request.json or {}
    label = b.get("label", "").strip()
    start = b.get("start_date", "")
    end   = b.get("end_date", "")
    if not (label and start and end):
        return err("label, start_date, end_date required")
    cur = run("INSERT INTO periods (label,start_date,end_date,is_active) VALUES (?,?,?,0)", (label, start, end))
    return jsonify({"id": cur.lastrowid, "label": label}), 201

@app.route("/api/periods/<int:pid>/activate", methods=["POST", "OPTIONS"])
@admin_required
@csrf_protect
def activate_period(pid):
    if request.method == "OPTIONS":
        return "", 204
    run("UPDATE periods SET is_active=0")
    run("UPDATE periods SET is_active=1 WHERE id=?", (pid,))
    return jsonify({"activated": pid})

@app.route("/api/periods/<int:pid>/close", methods=["POST", "OPTIONS"])
@admin_required
@csrf_protect
def close_period(pid):
    """Mark a period as closed and auto-advance `is_active` to the next
    unclosed month. Response includes the newly-active period so the
    frontend can jump to it without a second round-trip."""
    if request.method == "OPTIONS":
        return "", 204
    _ensure_period_close_columns()
    row = q1("SELECT id, label FROM periods WHERE id=?", (pid,))
    if not row:
        return err("Period not found", 404)
    u = get_current_user()
    uid = u["id"] if u else None
    run("UPDATE periods SET is_closed=1, closed_at=CURRENT_TIMESTAMP, closed_by=? WHERE id=?", (uid, pid))
    new_active_id = _activate_current_close_period()
    new_active = q1("SELECT * FROM periods WHERE id=?", (new_active_id,)) if new_active_id else None
    return jsonify({
        "closed": pid,
        "closed_label": row["label"],
        "active": dict(new_active) if new_active else None,
    })

@app.route("/api/periods/<int:pid>/reopen", methods=["POST", "OPTIONS"])
@admin_required
@csrf_protect
def reopen_period(pid):
    """Undo a close. If the reopened period is earlier than the currently
    active one, it becomes active again (it's now the earliest unclosed)."""
    if request.method == "OPTIONS":
        return "", 204
    _ensure_period_close_columns()
    row = q1("SELECT id, label FROM periods WHERE id=?", (pid,))
    if not row:
        return err("Period not found", 404)
    run("UPDATE periods SET is_closed=0, closed_at=NULL, closed_by=NULL WHERE id=?", (pid,))
    new_active_id = _activate_current_close_period()
    new_active = q1("SELECT * FROM periods WHERE id=?", (new_active_id,)) if new_active_id else None
    return jsonify({
        "reopened": pid,
        "reopened_label": row["label"],
        "active": dict(new_active) if new_active else None,
    })

# ── Users ─────────────────────────────────────────────────────────────────────

@app.route("/api/users", methods=["GET", "OPTIONS"])
@login_required
def get_users():
    if request.method == "OPTIONS":
        return "", 204
    return jsonify(rows_to_list(q("SELECT id,name,initials,email,role,color FROM users")))

@app.route("/api/users", methods=["POST", "OPTIONS"])
@admin_required
@csrf_protect
def create_user():
    if request.method == "OPTIONS":
        return "", 204
    b = request.json or {}
    if not all(k in b for k in ("name", "initials", "role", "color")):
        return err("name, initials, role, color required")
    cur = run(
        "INSERT INTO users (name,initials,email,role,color) VALUES (?,?,?,?,?)",
        (b["name"], b["initials"], b.get("email", ""), b["role"], b["color"]))
    return jsonify({"id": cur.lastrowid}), 201

@app.route("/api/users/<int:uid>", methods=["PATCH", "DELETE", "OPTIONS"])
@admin_required
@csrf_protect
def manage_user(uid):
    if request.method == "OPTIONS":
        return "", 204
    if request.method == "DELETE":
        if uid == DEFAULT_USER_ID:
            return err("Cannot delete the default admin user")
        tasks = q1("SELECT COUNT(*) AS n FROM tasks WHERE assignee_id=? OR reviewer_id=?", (uid, uid))["n"]
        recons = q1("SELECT COUNT(*) AS n FROM reconciliations WHERE assignee_id=?", (uid,))["n"]
        if tasks or recons:
            parts = []
            if tasks:  parts.append(f"{tasks} task(s)")
            if recons: parts.append(f"{recons} reconciliation(s)")
            return err(f"Cannot delete: user is still assigned to {' and '.join(parts)}. Reassign them first.", 409)
        run("DELETE FROM task_activity WHERE user_id=?", (uid,))
        run("DELETE FROM users WHERE id=?", (uid,))
        return jsonify({"deleted": uid})
    b = request.json or {}
    allowed = {"name", "initials", "email", "role", "color"}
    updates = {k: v for k, v in b.items() if k in allowed}
    if not updates:
        return err("No valid fields")
    safe_update("users", "id", allowed, updates, uid)
    return jsonify({"updated": uid})

# ── Categories ────────────────────────────────────────────────────────────────

@app.route("/api/categories", methods=["GET", "OPTIONS"])
@login_required
def get_categories():
    if request.method == "OPTIONS":
        return "", 204
    return jsonify(rows_to_list(q("SELECT * FROM categories ORDER BY sort_order")))

@app.route("/api/categories", methods=["POST", "OPTIONS"])
@admin_required
@csrf_protect
def create_category():
    if request.method == "OPTIONS":
        return "", 204
    b = request.json or {}
    name = b.get("name", "").strip()
    if not name:
        return err("name required")
    cur = run("INSERT INTO categories (name,sort_order) VALUES (?,?)", (name, b.get("sort_order", 99)))
    return jsonify({"id": cur.lastrowid, "name": name}), 201

@app.route("/api/categories/<int:cid>", methods=["PATCH", "DELETE", "OPTIONS"])
@admin_required
@csrf_protect
def manage_category(cid):
    if request.method == "OPTIONS":
        return "", 204
    if request.method == "DELETE":
        in_use = q1("SELECT COUNT(*) AS n FROM tasks WHERE category_id=?", (cid,))["n"]
        if in_use:
            return err(f"Cannot delete: category still has {in_use} task(s). Reassign or delete them first.", 409)
        run("DELETE FROM categories WHERE id=?", (cid,))
        return jsonify({"deleted": cid})
    b = request.json or {}
    allowed = {"name", "sort_order"}
    updates = {k: v for k, v in b.items() if k in allowed}
    if not updates:
        return err("No valid fields")
    safe_update("categories", "id", allowed, updates, cid)
    return jsonify({"updated": cid})

# ── Tasks ─────────────────────────────────────────────────────────────────────

TASK_SELECT = """
    SELECT t.*, c.name AS category_name,
           u1.name AS assignee_name, u1.initials AS assignee_initials, u1.color AS assignee_color,
           u2.name AS reviewer_name, u2.initials AS reviewer_initials,
           u3.name AS submitter_name
    FROM tasks t
    JOIN categories c ON c.id = t.category_id
    JOIN users u1 ON u1.id = t.assignee_id
    JOIN users u2 ON u2.id = t.reviewer_id
    LEFT JOIN users u3 ON u3.id = t.submitted_by
"""

@app.route("/api/tasks", methods=["GET", "OPTIONS"])
@login_required
def get_tasks():
    if request.method == "OPTIONS":
        return "", 204
    period_id = request.args.get("period_id")
    if not period_id:
        active = q1("SELECT id FROM periods WHERE is_active=1")
        period_id = active["id"] if active else None
    if not period_id:
        return jsonify([])
    return jsonify(rows_to_list(q(TASK_SELECT + " WHERE t.period_id=? ORDER BY c.sort_order, t.id", (period_id,))))

@app.route("/api/tasks", methods=["POST", "OPTIONS"])
@admin_required
@csrf_protect
def create_task():
    if request.method == "OPTIONS":
        return "", 204
    b = request.json or {}
    if not all(k in b for k in ("period_id", "category_id", "name", "assignee_id", "reviewer_id")):
        return err("Missing required fields")
    cur = run(
        "INSERT INTO tasks (period_id,category_id,name,assignee_id,reviewer_id,due_date,notes) VALUES (?,?,?,?,?,?,?)",
        (b["period_id"], b["category_id"], b["name"], b["assignee_id"], b["reviewer_id"],
         b.get("due_date"), b.get("notes", "")))
    return jsonify({"id": cur.lastrowid}), 201

@app.route("/api/tasks/<int:tid>", methods=["GET", "PATCH", "DELETE", "OPTIONS"])
@login_required
@csrf_protect
def manage_task(tid):
    if request.method == "OPTIONS":
        return "", 204
    if request.method == "GET":
        row = q1(TASK_SELECT + " WHERE t.id=?", (tid,))
        return jsonify(dict(row)) if row else err("Task not found", 404)
    if request.method == "DELETE":
        user = get_current_user()
        if user["role"] != "admin":
            return err("Admin required", 403)
        run("DELETE FROM task_activity WHERE task_id=?", (tid,))
        run("DELETE FROM tasks WHERE id=?", (tid,))
        return jsonify({"deleted": tid})
    # PATCH
    b = request.json or {}
    user = get_current_user()
    old = q1("SELECT * FROM tasks WHERE id=?", (tid,))
    if not old:
        return err("Task not found", 404)
    if user["role"] == "admin":
        allowed = {"status", "review_status", "notes", "assignee_id", "reviewer_id", "due_date", "name", "category_id", "recurrence"}
    else:
        if old["assignee_id"] != user["id"] and old["reviewer_id"] != user["id"]:
            return err("You can only update your own tasks", 403)
        allowed = {"status", "review_status", "notes", "recurrence"}
    updates = {k: v for k, v in b.items() if k in allowed}
    if not updates:
        return err("No valid fields")
    if updates.get("status") == "complete" and old["status"] != "complete":
        updates["completed_at"] = datetime.now(timezone.utc).isoformat()
    if updates.get("review_status") == "approved" and old["review_status"] != "approved":
        updates["approved_at"] = datetime.now(timezone.utc).isoformat()
    if updates.get("status") == "submitted" and (old["status"] if old else "") != "submitted":
        updates["submitted_at"] = datetime.now(timezone.utc).isoformat()
        updates["submitted_by"] = user["id"]
    safe_update("tasks", "id", allowed | {"completed_at", "approved_at", "submitted_at", "submitted_by"}, updates, tid)
    actor_id = user["id"]
    if "status" in updates:
        run("INSERT INTO task_activity (task_id,user_id,action,old_value,new_value) VALUES (?,?,?,?,?)",
            (tid, actor_id, "status_change", old["status"], updates["status"]))
    if "review_status" in updates:
        run("INSERT INTO task_activity (task_id,user_id,action,old_value,new_value) VALUES (?,?,?,?,?)",
            (tid, actor_id, "review_change", old["review_status"], updates["review_status"]))
    if "notes" in updates:
        run("INSERT INTO task_activity (task_id,user_id,action,old_value,new_value) VALUES (?,?,?,?,?)",
            (tid, actor_id, "note", None, updates["notes"]))
    return jsonify(dict(q1("SELECT * FROM tasks WHERE id=?", (tid,))))

# ── Reconciliations ───────────────────────────────────────────────────────────

RECON_SELECT = """
    SELECT r.*, u.name AS assignee_name, u.initials AS assignee_initials, u.color AS assignee_color,
           CASE WHEN r.expected_balance IS NOT NULL
                THEN r.qb_balance - r.expected_balance
                ELSE NULL END AS variance
    FROM reconciliations r
    JOIN users u ON u.id = r.assignee_id
"""

@app.route("/api/reconciliations", methods=["GET", "OPTIONS"])
@login_required
def get_reconciliations():
    if request.method == "OPTIONS":
        return "", 204
    period_id = request.args.get("period_id")
    if not period_id:
        active = q1("SELECT id FROM periods WHERE is_active=1")
        period_id = active["id"] if active else None
    if not period_id:
        return jsonify([])
    return jsonify(rows_to_list(q(RECON_SELECT + " WHERE r.period_id=? ORDER BY r.account_number", (period_id,))))

@app.route("/api/reconciliations", methods=["POST", "OPTIONS"])
@admin_required
@csrf_protect
def create_reconciliation():
    if request.method == "OPTIONS":
        return "", 204
    b = request.json or {}
    if not all(k in b for k in ("period_id", "account_number", "account_name", "assignee_id")):
        return err("Missing required fields")
    cur = run(
        "INSERT INTO reconciliations (period_id,account_number,account_name,assignee_id,qb_balance,expected_balance,status) VALUES (?,?,?,?,?,?,?)",
        (b["period_id"], b["account_number"], b["account_name"], b["assignee_id"],
         b.get("qb_balance"), b.get("expected_balance"), "open"))
    return jsonify({"id": cur.lastrowid}), 201

@app.route("/api/reconciliations/<int:rid>", methods=["PATCH", "DELETE", "OPTIONS"])
@login_required
@csrf_protect
def manage_reconciliation(rid):
    if request.method == "OPTIONS":
        return "", 204
    if request.method == "DELETE":
        user = get_current_user()
        if user["role"] != "admin":
            return err("Admin required", 403)
        run("DELETE FROM reconciliations WHERE id=?", (rid,))
        return jsonify({"deleted": rid})
    b = request.json or {}
    user = get_current_user()
    allowed = {"expected_balance", "status", "assignee_id", "account_number", "account_name"} if user["role"] == "admin" else {"expected_balance", "status"}
    updates = {k: v for k, v in b.items() if k in allowed}
    if not updates:
        return err("No valid fields")
    updates["last_updated_at"] = datetime.now(timezone.utc).isoformat()
    safe_update("reconciliations", "id", allowed | {"last_updated_at"}, updates, rid)
    return jsonify(dict(q1(RECON_SELECT + " WHERE r.id=?", (rid,))))

# ── Dashboard ─────────────────────────────────────────────────────────────────

@app.route("/api/dashboard", methods=["GET", "OPTIONS"])
@login_required
def dashboard():
    if request.method == "OPTIONS":
        return "", 204
    period_id = request.args.get("period_id")
    if not period_id:
        active = q1("SELECT id FROM periods WHERE is_active=1")
        period_id = active["id"] if active else None
    if not period_id:
        return err("No active period", 404)
    tasks_all  = rows_to_list(q("SELECT status,review_status,assignee_id,category_id FROM tasks WHERE period_id=?", (period_id,)))
    recons_all = rows_to_list(q("SELECT status FROM reconciliations WHERE period_id=?", (period_id,)))
    total      = len(tasks_all)
    complete   = sum(1 for t in tasks_all if t["status"] == "complete")
    approved   = sum(1 for t in tasks_all if t["review_status"] == "approved")
    recon_done = sum(1 for r in recons_all if r["status"] == "reconciled")
    users = rows_to_list(q("SELECT id,name,initials,color FROM users"))
    for u in users:
        ut = [t for t in tasks_all if t["assignee_id"] == u["id"]]
        u["tasks_total"]    = len(ut)
        u["tasks_complete"] = sum(1 for t in ut if t["status"] == "complete")
    cats = rows_to_list(q("SELECT id,name FROM categories ORDER BY sort_order"))
    for c in cats:
        ct = [t for t in tasks_all if t["category_id"] == c["id"]]
        c["tasks_total"]    = len(ct)
        c["tasks_complete"] = sum(1 for t in ct if t["status"] == "complete")
    attention = rows_to_list(q("""
        SELECT t.id, t.name, t.status, t.review_status, t.due_date,
               u.name AS assignee_name, u.color AS assignee_color, c.name AS category_name
        FROM tasks t
        JOIN users u ON u.id = t.assignee_id
        JOIN categories c ON c.id = t.category_id
        WHERE t.period_id=? AND (t.review_status='needs_revision' OR t.status='open')
        ORDER BY t.due_date LIMIT 8""", (period_id,)))
    # open issues count
    try:
        open_issues = q1("SELECT COUNT(*) AS n FROM open_items WHERE period_id=? AND status='open'", (period_id,))
        open_issues_count = open_issues["n"] if open_issues else 0
    except Exception:
        open_issues_count = 0
    # avg days to close
    try:
        _ensure_period_close_columns()
        closed_rows = q("SELECT end_date, closed_at FROM periods WHERE is_closed=1 AND closed_at IS NOT NULL")
        days_list = []
        for cr in closed_rows:
            try:
                from datetime import datetime as _dt2
                d1 = _dt2.fromisoformat(cr["end_date"][:10])
                d2 = _dt2.fromisoformat(cr["closed_at"][:19].replace("T", " ").split(".")[0])
                days_list.append(max(0, (d2.date() - d1.date()).days))
            except Exception:
                pass
        avg_days_to_close = round(sum(days_list) / len(days_list), 1) if days_list else None
    except Exception:
        avg_days_to_close = None
    return jsonify({
        "period_id":        period_id,
        "tasks_total":      total,
        "tasks_complete":   complete,
        "tasks_approved":   approved,
        "close_pct":        round(complete / total * 100) if total else 0,
        "approval_pct":     round(approved / total * 100) if total else 0,
        "recon_total":      len(recons_all),
        "recon_done":       recon_done,
        "recon_pct":        round(recon_done / len(recons_all) * 100) if recons_all else 0,
        "by_user":          users,
        "by_category":      cats,
        "attention":        attention,
        "open_issues":      open_issues_count,
        "avg_days_to_close": avg_days_to_close,
    })

# ── Feature 2: Task Dependencies ─────────────────────────────────────────────

@app.route("/api/tasks/<int:tid>/dependencies", methods=["GET", "OPTIONS"])
@login_required
def get_task_dependencies(tid):
    if request.method == "OPTIONS":
        return "", 204
    _ensure_new_features_schema()
    rows = q("""SELECT t.id, t.name, t.status, t.category_id, c.name AS category_name
                FROM task_dependencies td
                JOIN tasks t ON t.id = td.depends_on_id
                JOIN categories c ON c.id = t.category_id
                WHERE td.task_id=?""", (tid,))
    return jsonify(rows_to_list(rows))

@app.route("/api/tasks/<int:tid>/dependencies", methods=["POST"])
@login_required
@csrf_protect
def add_task_dependency(tid):
    _ensure_new_features_schema()
    b = request.get_json(silent=True) or {}
    dep_id = b.get("depends_on_id")
    if not dep_id:
        return err("depends_on_id required")
    if dep_id == tid:
        return err("Task cannot depend on itself")
    try:
        run("INSERT OR IGNORE INTO task_dependencies (task_id, depends_on_id) VALUES (?,?)", (tid, dep_id))
    except Exception as e:
        return err(str(e))
    return jsonify({"ok": True}), 201

@app.route("/api/tasks/<int:tid>/dependencies/<int:dep_id>", methods=["DELETE", "OPTIONS"])
@login_required
@csrf_protect
def remove_task_dependency(tid, dep_id):
    if request.method == "OPTIONS":
        return "", 204
    _ensure_new_features_schema()
    run("DELETE FROM task_dependencies WHERE task_id=? AND depends_on_id=?", (tid, dep_id))
    return jsonify({"ok": True})

# ── Feature 3: Roll-forward ────────────────────────────────────────────────────

@app.route("/api/periods/<int:pid>/rollforward", methods=["POST", "OPTIONS"])
@admin_required
@csrf_protect
def rollforward_period(pid):
    if request.method == "OPTIONS":
        return "", 204
    b = request.get_json(silent=True) or {}
    target_id = b.get("target_period_id")
    if not target_id:
        return err("target_period_id required")
    src_tasks = q("SELECT * FROM tasks WHERE period_id=?", (pid,))
    db = get_db()
    copied = 0
    for t in src_tasks:
        if t["recurrence"] in (None, "none", "") and not b.get("include_all"):
            pass  # include all tasks regardless (simple roll-forward)
        db.execute(
            """INSERT INTO tasks (period_id, category_id, name, assignee_id, reviewer_id, due_date,
               status, review_status, notes, recurrence)
               VALUES (?,?,?,?,?,?,?,?,?,?)""",
            (target_id, t["category_id"], t["name"], t["assignee_id"], t["reviewer_id"],
             t["due_date"], "open", "pending", "", t["recurrence"] or "none"))
        copied += 1
    db.commit()
    return jsonify({"copied": copied})

# ── Feature 4: Checklist Templates ────────────────────────────────────────────

@app.route("/api/templates", methods=["GET", "OPTIONS"])
@login_required
def get_templates():
    if request.method == "OPTIONS":
        return "", 204
    _ensure_new_features_schema()
    rows = q("SELECT * FROM checklist_templates ORDER BY created_at DESC")
    return jsonify(rows_to_list(rows))

@app.route("/api/templates", methods=["POST"])
@admin_required
@csrf_protect
def create_template():
    _ensure_new_features_schema()
    b = request.get_json(silent=True) or {}
    name = (b.get("name") or "").strip()
    if not name:
        return err("name required")
    u = get_current_user()
    cur = run("INSERT INTO checklist_templates (name, period_type, description, created_by) VALUES (?,?,?,?)",
              (name, b.get("period_type", "monthly"), b.get("description", ""), u["id"] if u else DEFAULT_USER_ID))
    return jsonify({"id": cur.lastrowid, "name": name}), 201

@app.route("/api/templates/<int:tmpl_id>", methods=["DELETE", "OPTIONS"])
@admin_required
@csrf_protect
def delete_template(tmpl_id):
    if request.method == "OPTIONS":
        return "", 204
    run("DELETE FROM template_tasks WHERE template_id=?", (tmpl_id,))
    run("DELETE FROM checklist_templates WHERE id=?", (tmpl_id,))
    return jsonify({"deleted": tmpl_id})

@app.route("/api/templates/<int:tmpl_id>/tasks", methods=["GET", "OPTIONS"])
@login_required
def get_template_tasks(tmpl_id):
    if request.method == "OPTIONS":
        return "", 204
    _ensure_new_features_schema()
    rows = q("SELECT * FROM template_tasks WHERE template_id=? ORDER BY sort_order, id", (tmpl_id,))
    return jsonify(rows_to_list(rows))

@app.route("/api/templates/<int:tmpl_id>/tasks", methods=["POST"])
@admin_required
@csrf_protect
def add_template_task(tmpl_id):
    _ensure_new_features_schema()
    b = request.get_json(silent=True) or {}
    title = (b.get("title") or "").strip()
    if not title:
        return err("title required")
    cur = run("INSERT INTO template_tasks (template_id, title, category_id, assignee_id, reviewer_id, day_target, sort_order) VALUES (?,?,?,?,?,?,?)",
              (tmpl_id, title, b.get("category_id"), b.get("assignee_id"), b.get("reviewer_id"),
               b.get("day_target"), b.get("sort_order", 0)))
    return jsonify({"id": cur.lastrowid}), 201

@app.route("/api/templates/<int:tmpl_id>/tasks/<int:ttid>", methods=["DELETE", "OPTIONS"])
@admin_required
@csrf_protect
def remove_template_task(tmpl_id, ttid):
    if request.method == "OPTIONS":
        return "", 204
    run("DELETE FROM template_tasks WHERE id=? AND template_id=?", (ttid, tmpl_id))
    return jsonify({"ok": True})

@app.route("/api/templates/<int:tmpl_id>/apply", methods=["POST", "OPTIONS"])
@admin_required
@csrf_protect
def apply_template(tmpl_id):
    if request.method == "OPTIONS":
        return "", 204
    _ensure_new_features_schema()
    b = request.get_json(silent=True) or {}
    period_id = b.get("period_id")
    if not period_id:
        return err("period_id required")
    period = q1("SELECT * FROM periods WHERE id=?", (period_id,))
    if not period:
        return err("Period not found", 404)
    tasks = q("SELECT * FROM template_tasks WHERE template_id=? ORDER BY sort_order, id", (tmpl_id,))
    db = get_db()
    created = 0
    for t in tasks:
        due = None
        if t["day_target"] and period["start_date"]:
            from datetime import date as _date, timedelta as _td
            start = _date.fromisoformat(period["start_date"])
            due = (start + _td(days=t["day_target"] - 1)).isoformat()
        db.execute("""INSERT INTO tasks (period_id, category_id, name, assignee_id, reviewer_id, due_date, status, review_status)
                      VALUES (?,?,?,?,?,?,?,?)""",
                   (period_id, t["category_id"], t["title"], t["assignee_id"], t["reviewer_id"],
                    due, "open", "pending"))
        created += 1
    db.commit()
    return jsonify({"created": created})

# ── Feature 6: Open Items / Issues Log ────────────────────────────────────────

@app.route("/api/open_items", methods=["GET", "OPTIONS"])
@login_required
def get_open_items():
    if request.method == "OPTIONS":
        return "", 204
    _ensure_new_features_schema()
    period_id = request.args.get("period_id")
    if not period_id:
        active = q1("SELECT id FROM periods WHERE is_active=1")
        period_id = active["id"] if active else None
    if not period_id:
        return jsonify([])
    rows = q("""SELECT oi.*, u1.name AS assigned_to_name, u2.name AS created_by_name
                FROM open_items oi
                LEFT JOIN users u1 ON u1.id = oi.assigned_to
                LEFT JOIN users u2 ON u2.id = oi.created_by
                WHERE oi.period_id=? ORDER BY oi.created_at DESC""", (period_id,))
    return jsonify(rows_to_list(rows))

@app.route("/api/open_items", methods=["POST"])
@login_required
@csrf_protect
def create_open_item():
    _ensure_new_features_schema()
    b = request.get_json(silent=True) or {}
    title = (b.get("title") or "").strip()
    period_id = b.get("period_id")
    if not title or not period_id:
        return err("title and period_id required")
    u = get_current_user()
    uid = u["id"] if u else DEFAULT_USER_ID
    cur = run("""INSERT INTO open_items (period_id, title, description, status, priority, assigned_to, created_by)
                 VALUES (?,?,?,?,?,?,?)""",
              (period_id, title, b.get("description", ""), b.get("status", "open"),
               b.get("priority", "medium"), b.get("assigned_to"), uid))
    run("INSERT INTO task_activity (task_id, user_id, action, new_value) VALUES (?,?,?,?)",
        (cur.lastrowid, uid, "created", f"open_item:{title}"))
    return jsonify({"id": cur.lastrowid}), 201

@app.route("/api/open_items/<int:oid>", methods=["PATCH", "DELETE", "OPTIONS"])
@login_required
@csrf_protect
def manage_open_item(oid):
    if request.method == "OPTIONS":
        return "", 204
    _ensure_new_features_schema()
    if request.method == "DELETE":
        u = get_current_user()
        if u["role"] != "admin":
            return err("Admin required", 403)
        run("DELETE FROM open_items WHERE id=?", (oid,))
        return jsonify({"deleted": oid})
    b = request.get_json(silent=True) or {}
    allowed = {"status", "description", "assigned_to", "priority", "resolved_at", "resolved_by"}
    updates = {k: v for k, v in b.items() if k in allowed}
    if not updates:
        return err("No valid fields")
    u = get_current_user()
    uid = u["id"] if u else DEFAULT_USER_ID
    if updates.get("status") == "resolved":
        updates.setdefault("resolved_at", datetime.now(timezone.utc).isoformat())
        updates.setdefault("resolved_by", uid)
    safe_update("open_items", "id", allowed, updates, oid)
    run("INSERT INTO task_activity (task_id, user_id, action, new_value) VALUES (?,?,?,?)",
        (oid, uid, "status_change", updates.get("status", "")))
    row = q1("""SELECT oi.*, u1.name AS assigned_to_name FROM open_items oi
                LEFT JOIN users u1 ON u1.id = oi.assigned_to WHERE oi.id=?""", (oid,))
    return jsonify(dict(row)) if row else err("Not found", 404)

# ── Feature 8: Attachments ─────────────────────────────────────────────────────

@app.route("/api/attachments", methods=["GET", "POST", "OPTIONS"])
@login_required
@csrf_protect
def attachments():
    if request.method == "OPTIONS":
        return "", 204
    _ensure_new_features_schema()
    if request.method == "GET":
        task_id = request.args.get("task_id")
        recon_id = request.args.get("recon_id")
        if task_id:
            rows = q("""SELECT a.id, a.filename, a.content_type, a.uploaded_at, u.name AS uploaded_by_name
                        FROM attachments a LEFT JOIN users u ON u.id=a.uploaded_by
                        WHERE a.task_id=? ORDER BY a.uploaded_at DESC""", (task_id,))
        elif recon_id:
            rows = q("""SELECT a.id, a.filename, a.content_type, a.uploaded_at, u.name AS uploaded_by_name
                        FROM attachments a LEFT JOIN users u ON u.id=a.uploaded_by
                        WHERE a.recon_id=? ORDER BY a.uploaded_at DESC""", (recon_id,))
        else:
            return err("task_id or recon_id required")
        return jsonify(rows_to_list(rows))
    # POST — file upload
    f = request.files.get("file")
    if not f:
        return err("file required")
    data = f.read()
    if len(data) > 5 * 1024 * 1024:
        return err("File too large (max 5 MB)")
    u = get_current_user()
    uid = u["id"] if u else DEFAULT_USER_ID
    task_id = request.form.get("task_id")
    recon_id = request.form.get("recon_id")
    cur = run("INSERT INTO attachments (task_id, recon_id, filename, content_type, data, uploaded_by) VALUES (?,?,?,?,?,?)",
              (task_id, recon_id, f.filename, f.content_type or "application/octet-stream", data, uid))
    return jsonify({"id": cur.lastrowid, "filename": f.filename}), 201

@app.route("/api/attachments/<int:att_id>/download", methods=["GET"])
@login_required
def download_attachment(att_id):
    _ensure_new_features_schema()
    row = q1("SELECT filename, content_type, data FROM attachments WHERE id=?", (att_id,))
    if not row:
        return err("Not found", 404)
    from flask import Response
    return Response(bytes(row["data"]), mimetype=row["content_type"] or "application/octet-stream",
                    headers={"Content-Disposition": f'attachment; filename="{row["filename"]}"'})

@app.route("/api/attachments/<int:att_id>", methods=["DELETE", "OPTIONS"])
@admin_required
@csrf_protect
def delete_attachment(att_id):
    if request.method == "OPTIONS":
        return "", 204
    _ensure_new_features_schema()
    run("DELETE FROM attachments WHERE id=?", (att_id,))
    return jsonify({"deleted": att_id})

# ── Feature 9: Journal Entries ─────────────────────────────────────────────────

@app.route("/api/journal_entries", methods=["GET", "OPTIONS"])
@login_required
def get_journal_entries():
    if request.method == "OPTIONS":
        return "", 204
    _ensure_new_features_schema()
    period_id = request.args.get("period_id")
    if not period_id:
        active = q1("SELECT id FROM periods WHERE is_active=1")
        period_id = active["id"] if active else None
    if not period_id:
        return jsonify([])
    rows = q("""SELECT je.*, u1.name AS preparer_name, u2.name AS reviewer_name
                FROM journal_entries je
                LEFT JOIN users u1 ON u1.id = je.preparer_id
                LEFT JOIN users u2 ON u2.id = je.reviewer_id
                WHERE je.period_id=? ORDER BY je.id DESC""", (period_id,))
    return jsonify(rows_to_list(rows))

@app.route("/api/journal_entries", methods=["POST"])
@login_required
@csrf_protect
def create_journal_entry():
    _ensure_new_features_schema()
    b = request.get_json(silent=True) or {}
    desc = (b.get("description") or "").strip()
    period_id = b.get("period_id")
    if not desc or not period_id:
        return err("description and period_id required")
    u = get_current_user()
    uid = u["id"] if u else DEFAULT_USER_ID
    cur = run("""INSERT INTO journal_entries (period_id, je_number, description, debit_account, credit_account,
                 amount, preparer_id, reviewer_id, status, prepared_at, notes)
                 VALUES (?,?,?,?,?,?,?,?,?,CURRENT_TIMESTAMP,?)""",
              (period_id, b.get("je_number", ""), desc, b.get("debit_account", ""),
               b.get("credit_account", ""), b.get("amount"), uid, b.get("reviewer_id"),
               b.get("status", "draft"), b.get("notes", "")))
    run("INSERT INTO task_activity (task_id, user_id, action, new_value) VALUES (?,?,?,?)",
        (cur.lastrowid, uid, "created", f"je:{desc}"))
    return jsonify({"id": cur.lastrowid}), 201

@app.route("/api/journal_entries/<int:jeid>", methods=["PATCH", "DELETE", "OPTIONS"])
@login_required
@csrf_protect
def manage_journal_entry(jeid):
    if request.method == "OPTIONS":
        return "", 204
    _ensure_new_features_schema()
    if request.method == "DELETE":
        u = get_current_user()
        if u["role"] != "admin":
            return err("Admin required", 403)
        run("DELETE FROM journal_entries WHERE id=?", (jeid,))
        return jsonify({"deleted": jeid})
    b = request.get_json(silent=True) or {}
    allowed = {"status", "notes", "reviewer_id", "je_number", "description",
               "debit_account", "credit_account", "amount", "approved_at", "submitted_at"}
    updates = {k: v for k, v in b.items() if k in allowed}
    if not updates:
        return err("No valid fields")
    u = get_current_user()
    uid = u["id"] if u else DEFAULT_USER_ID
    old = q1("SELECT status FROM journal_entries WHERE id=?", (jeid,))
    if updates.get("status") == "submitted" and (old["status"] if old else "") != "submitted":
        updates["submitted_at"] = datetime.now(timezone.utc).isoformat()
    if updates.get("status") == "approved" and (old["status"] if old else "") != "approved":
        updates["approved_at"] = datetime.now(timezone.utc).isoformat()
    safe_update("journal_entries", "id", allowed, updates, jeid)
    run("INSERT INTO task_activity (task_id, user_id, action, old_value, new_value) VALUES (?,?,?,?,?)",
        (jeid, uid, "status_change", old["status"] if old else None, updates.get("status", "")))
    row = q1("""SELECT je.*, u1.name AS preparer_name, u2.name AS reviewer_name
                FROM journal_entries je
                LEFT JOIN users u1 ON u1.id = je.preparer_id
                LEFT JOIN users u2 ON u2.id = je.reviewer_id
                WHERE je.id=?""", (jeid,))
    return jsonify(dict(row)) if row else err("Not found", 404)

# ── Feature 10: Task Activity ──────────────────────────────────────────────────

@app.route("/api/tasks/<int:tid>/activity", methods=["GET", "OPTIONS"])
@login_required
def get_task_activity(tid):
    if request.method == "OPTIONS":
        return "", 204
    rows = q("""SELECT ta.*, u.name AS user_name, u.initials AS user_initials
                FROM task_activity ta
                LEFT JOIN users u ON u.id = ta.user_id
                WHERE ta.task_id=? ORDER BY ta.created_at ASC""", (tid,))
    return jsonify(rows_to_list(rows))

# ── Feature 12/13: Analytics / Time-to-Close ──────────────────────────────────

@app.route("/api/analytics/time_to_close", methods=["GET", "OPTIONS"])
@login_required
def time_to_close():
    if request.method == "OPTIONS":
        return "", 204
    _ensure_period_close_columns()
    rows = q("""SELECT id, label, end_date, closed_at
                FROM periods WHERE is_closed=1 AND closed_at IS NOT NULL
                ORDER BY end_date ASC""")
    result = []
    for r in rows:
        try:
            end = r["end_date"]
            closed = r["closed_at"]
            if end and closed:
                from datetime import datetime as _dt
                d_end = _dt.fromisoformat(end[:10])
                d_closed = _dt.fromisoformat(closed[:19].replace("T", " ").split(".")[0])
                days = max(0, (d_closed.date() - d_end.date()).days)
            else:
                days = None
        except Exception:
            days = None
        result.append({"id": r["id"], "label": r["label"], "end_date": r["end_date"],
                        "closed_at": r["closed_at"], "days_to_close": days})
    return jsonify(result)

# ── Frontend ──────────────────────────────────────────────────────────────────

@app.route("/")
def index():
    return send_from_directory(STATIC_DIR, "index.html")

@app.route("/<path:filename>")
def static_files(filename):
    filepath = os.path.join(STATIC_DIR, filename)
    if os.path.exists(filepath):
        return send_from_directory(STATIC_DIR, filename)
    return send_from_directory(STATIC_DIR, "index.html")

@app.errorhandler(404)
def not_found(e):
    return jsonify({"error": "Not found"}), 404

@app.errorhandler(500)
def server_error(e):
    traceback.print_exc()
    return jsonify({"error": "Internal server error"}), 500

if __name__ == "__main__":
    from init_db import init
    init()
    app.run(debug=True, port=5000, use_reloader=False)